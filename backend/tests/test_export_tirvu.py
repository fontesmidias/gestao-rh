"""Testes de CONTEÚDO do export do Tirvu — sem banco, sem containers.

A raiz dos feedbacks de 2026-07-24 foi um export que RODAVA mas cujas células
saíam com o valor errado (texto em vez de ID; CTPS série "0000"). O smoke valida
que o arquivo abre; estes testes validam o que vai DENTRO das células.

Rode: PYTHONPATH=. .venv/Scripts/python.exe tests/test_export_tirvu.py
"""

from openpyxl import load_workbook
import io

from app.services import export_tirvu as t


# ---- CTPS: número = 7 primeiros dígitos do CPF, série = 4 últimos ----
assert t.ctps_do_cpf("123.456.789-09") == ("1234567", "8909"), t.ctps_do_cpf("12345678909")
assert t.ctps_do_cpf("00000000000") == ("0000000", "0000")
assert t.ctps_do_cpf("123") == ("", "")  # CPF inválido não deriva CTPS
assert t.ctps_do_cpf(None) == ("", "")


# ---- normalizador de cargo: colapsa espaço, desacentua, minúsculo ----
assert t.normalizar_cargo("  Analista  DF  Jr ") == "analista df jr"
assert t.normalizar_cargo("Vigía") == t.normalizar_cargo("vigia") == "vigia"
assert t.normalizar_cargo(None) == ""


# ---- linha_tirvu escreve os IDs do Tirvu, não os textos ----
# stubs mínimos: só o que linha_tirvu lê de cada objeto.
class _Stub:
    def __init__(self, **kw):
        self.__dict__.update(kw)


class _DBFake:
    """Devolve objetos por (modelo, id) de um dicionário pré-carregado e resolve
    o de-para de cargo por scalar()."""
    def __init__(self, objetos, cargo_id=None):
        self._objetos = objetos
        self._cargo_id = cargo_id

    def get(self, modelo, ident):
        return self._objetos.get((modelo.__name__, ident))

    def scalar(self, _stmt):
        # usado só por tirvu_id_do_cargo — devolvemos o de-para simulado
        if self._cargo_id is None:
            return None
        return _Stub(tirvu_id=self._cargo_id)


from app.models.candidato import Candidato, Empresa, Jornada, PostoServico
from app.models.ficha import DadosPessoais, DocumentosIdentificacao, Endereco

cid = "cand-1"
cand = _Stub(id=cid, nome_completo="Fulano de Tal", cpf="123.456.789-09",
             cargo_funcao="Analista DF Jr", posto_servico_id="p1", empresa_id="e1",
             jornada_id="j1", registra_ponto=True, celular_whatsapp="(61) 99999-8888",
             salario_base="R$ 2.000,00", matricula="0001234", data_admissao="01/02/2026",
             data_nascimento="10/10/1990")
objetos = {
    ("Candidato", cid): cand,
    ("DadosPessoais", cid): _Stub(sexo=_Stub(value="masculino"), data_nascimento="10/10/1990"),
    ("Endereco", cid): _Stub(logradouro="Rua X", numero="10", complemento="",
                             logradouro_numero_complemento=None, cep="70000-000",
                             bairro="Centro", cidade="Brasília", uf="DF"),
    ("DocumentosIdentificacao", cid): _Stub(cpf="123.456.789-09", ctps_numero="12345678909",
                                            ctps_serie="0000", pis_nis_pasep="12345678901"),
    ("PostoServico", "p1"): _Stub(tirvu_id="49", nome="GHS"),
    ("Empresa", "e1"): _Stub(tirvu_id="1", razao_social="GREEN HOUSE LTDA"),
    ("Jornada", "j1"): _Stub(tirvu_id="246", descricao="GHS SEDE - 2A A 5A ..."),
}
db = _DBFake(objetos, cargo_id="50")

linha = t.linha_tirvu(db, cand, gerar_matricula=False)

# Empresa/Posto/Cargo/Jornada saem como ID NUMÉRICO — não o texto (bug 2026-07-24)
assert linha["Empresa"] == "1", linha["Empresa"]
assert linha["Posto de Serviço"] == "49", linha["Posto de Serviço"]
assert linha["Cargo"] == "50", linha["Cargo"]
assert linha["Descrição da Jornada de Trabalho"] == "246", linha["Descrição da Jornada de Trabalho"]
# CTPS derivada do CPF mesmo com "0000" gravado no banco (export re-deriva)
assert linha["CTPS Número"] == "1234567", linha["CTPS Número"]
assert linha["CTPS Série"] == "8909", linha["CTPS Série"]
# CEP com hífen, cidade preservada com acento, PIS sem máscara
assert linha["Endereço - CEP"] == "70000-000", linha["Endereço - CEP"]
assert linha["Endereço - Cidade"] == "Brasília", linha["Endereço - Cidade"]
assert linha["PIS"] == "12345678901"


# ---- CTPS: sem CPF, cai no gravado (fallback do elif) ----
cand_sem_cpf = _Stub(**{**cand.__dict__, "cpf": ""})
objs_sem_cpf = {
    **objetos,
    ("Candidato", cid): cand_sem_cpf,
    ("DocumentosIdentificacao", cid): _Stub(cpf="", ctps_numero="9998",
                                            ctps_serie="12", pis_nis_pasep="12345678901"),
}
linha_sc = t.linha_tirvu(_DBFake(objs_sem_cpf, cargo_id="50"), cand_sem_cpf,
                         gerar_matricula=False)
assert linha_sc["CTPS Número"] == "9998", linha_sc["CTPS Número"]
assert linha_sc["CTPS Série"] == "12", linha_sc["CTPS Série"]

# CPF INVÁLIDO (não-vazio, ≠11 dígitos) também cai no gravado (B1 da revisão)
cand_cpf_sujo = _Stub(**{**cand.__dict__, "cpf": "123"})
objs_sujo = {
    **objetos,
    ("Candidato", cid): cand_cpf_sujo,
    ("DocumentosIdentificacao", cid): _Stub(cpf="123", ctps_numero="7777",
                                            ctps_serie="55", pis_nis_pasep="1"),
}
linha_sj = t.linha_tirvu(_DBFake(objs_sujo, cargo_id="50"), cand_cpf_sujo,
                         gerar_matricula=False)
assert linha_sj["CTPS Número"] == "7777", linha_sj["CTPS Número"]


# ---- pendência quando o ID do Tirvu falta ----
db_sem_ids = _DBFake({**objetos,
                      ("Empresa", "e1"): _Stub(tirvu_id=None, razao_social="X"),
                      ("Jornada", "j1"): _Stub(tirvu_id=None, descricao="Y"),
                      ("PostoServico", "p1"): _Stub(tirvu_id=None, nome="Z")},
                     cargo_id=None)
linha2 = t.linha_tirvu(db_sem_ids, cand, gerar_matricula=False)
pend = t.pendencias_linha(linha2)
assert "ID Tirvu da empresa" in pend, pend
assert "ID Tirvu do posto" in pend, pend
assert "ID Tirvu do cargo" in pend, pend
assert "ID Tirvu da jornada" in pend, pend


# ---- workbook: aba Plan1, célula do cargo contém o ID ----
wb_bytes = t.montar_workbook_tirvu([linha])
wb = load_workbook(io.BytesIO(wb_bytes))
ws = wb.active
assert ws.title == "Plan1", ws.title
cabecalho = [c.value for c in ws[1]]
i_cargo = cabecalho.index("Cargo")
assert ws.cell(row=2, column=i_cargo + 1).value == "50"

print("test_export_tirvu: OK")
