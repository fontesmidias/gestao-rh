"""A planilha de uniforme que vai ANEXA no aviso (v2.81).

Pedido do Bruno (2026-08-07):

    "No corpo do email de envio de uniformes, tem que ir os dados da pessoa,
     como nome, CPF, cargo, posto e medidas. (…) quero algo que seja possível
     através da leitura do email, os responsáveis do uniforme identificarem as
     informações, sem a necessidade de entrar no sistema."

⚠️ **Isto REVERTE a decisão da v2.07**, e a reversão é consciente. Lá ele pediu
a mesma coisa e, perguntado, escolheu o contrário — dados só na tela, e-mail
como empurrão. O uso mostrou o custo: quem compra e separa uniforme não é
usuário do painel, e obrigá-lo a entrar para ver três medidas transformava um
recado em tarefa. Perguntado de novo (2026-08-08), escolheu a planilha ANEXA.

O que este teste trava:

1. **As 7 colunas, na ordem.** Quem abre precisa achar sempre a mesma coluna no
   mesmo lugar. A ordem é explícita no serviço, não a união das chaves do dict.
2. **O CPF sai MASCARADO e vem da ficha**, não do convite: o da ficha é o que a
   pessoa digitou e foi conferido contra o documento.
3. **Só o que serve para o uniforme.** Endereço, banco, salário e PIS estão a um
   `getattr` de distância e NÃO entram — anexo circula, e o que não é necessário
   para a tarefa não deve viajar junto (minimização, regra do
   `curriculo_texto.py`).
4. **Falha ao montar o anexo não segura o aviso.** O e-mail sai sem ele, com o
   link da tela — que é como funcionava antes. Perder o aviso inteiro por causa
   de um .xlsx seria trocar um problema pequeno por um maior.
5. **O texto do template diz que há anexo.** Ele afirmava o contrário ("não vai
   por e-mail") até a v2.80; instrução que o sistema não cumpre é a armadilha
   da v2.74.

Mutações verificadas:
  1. CPF sai cru (sem máscara)                 -> bloco 2 falha
  2. ordem das colunas vira união das chaves   -> bloco 1 falha
  3. falha do anexo derruba o aviso            -> bloco 4 falha
  4. CPF do convite vence o da ficha           -> bloco 2 falha

Precisa dos containers de teste.

Rode: PYTHONPATH=. .venv/Scripts/python.exe tests/test_uniforme_planilha.py
"""

import io
import os
import uuid

for _chave, _valor in dict(
    DATABASE_URL="postgresql+psycopg://admissao:admissao@localhost:55432/admissao",
    MINIO_ENDPOINT="localhost:59000",
    MINIO_ACCESS_KEY="minio",
    MINIO_SECRET_KEY="minio12345",
    MINIO_SECURE="false",
    RH_ADMIN_EMAIL="rh@greenhousedf.com.br",
    RH_ADMIN_PASSWORD="senha-teste-123",
    SECRET_KEY="segredo-de-teste",
    BASE_URL="http://localhost:8090",
).items():
    os.environ.setdefault(_chave, _valor)

from openpyxl import load_workbook  # noqa: E402

import app.main  # noqa: E402,F401  (registra os modelos; resolve as FKs)
from app.core.db import SessionLocal  # noqa: E402
from app.models.candidato import Candidato, PostoServico  # noqa: E402
from app.models.ficha import (DadosProfissionaisBancarios,  # noqa: E402
                              DocumentosIdentificacao)
from app.services.email import _tipo_do_anexo  # noqa: E402
from app.services.uniforme_planilha import (COLUNAS, linha_uniforme,  # noqa: E402
                                            montar_planilha, nome_do_arquivo)

SUF = uuid.uuid4().hex[:8]
falhas: list[str] = []


def checar(cond, msg):
    if cond:
        print(f"  ok   {msg}")
    else:
        print(f"  FALHOU  {msg}")
        falhas.append(msg)


def criar_pessoa(*, cpf_ficha=None, cpf_convite=None, com_medidas=True):
    """Cria a pessoa com posto, medidas e CPF — devolve o id."""
    with SessionLocal() as db:
        posto = PostoServico(nome=f"INEP - {SUF} - PORTARIA {uuid.uuid4().hex[:4]}",
                             sigla=f"S{uuid.uuid4().hex[:6]}")
        db.add(posto)
        db.flush()
        cand = Candidato(nome_completo=f"Maria Uniforme {uuid.uuid4().hex[:4]}",
                         email=f"unif-{uuid.uuid4().hex[:8]}@exemplo.com",
                         cargo_funcao="Recepcionista", posto_servico_id=posto.id,
                         cpf=cpf_convite)
        db.add(cand)
        db.flush()
        if com_medidas:
            db.add(DadosProfissionaisBancarios(
                candidato_id=cand.id, tamanho_calca="42",
                tamanho_camisa="M", tamanho_calcado="38",
                # Campos que NÃO devem sair na planilha — ver bloco 3.
                banco="Banco do Brasil", pix_tipo="cpf", pix_chave="12345678909"))
        if cpf_ficha:
            db.add(DocumentosIdentificacao(candidato_id=cand.id, cpf=cpf_ficha))
        db.commit()
        return str(cand.id), posto.nome


# --------------------------------------------------------------------------
print("\n1. as 7 colunas, na ORDEM")
# --------------------------------------------------------------------------
# ⚠️ Mutação 2: montar pela união das chaves -> a ordem muda e isto falha.
cid, nome_posto = criar_pessoa(cpf_ficha="12345678909")
with SessionLocal() as db:
    cand = db.get(Candidato, uuid.UUID(cid))
    bytes_xlsx = montar_planilha(db, [cand])

checar(len(bytes_xlsx) > 1000, f"a planilha tem conteúdo ({len(bytes_xlsx)} bytes)")
ws = load_workbook(io.BytesIO(bytes_xlsx)).active
cabecalho = [c.value for c in ws[1]]
checar(cabecalho == list(COLUNAS),
       f"cabeçalho na ordem exata do serviço — quem abre acha sempre a mesma "
       f"coluna no mesmo lugar (veio {cabecalho})")
checar(len(cabecalho) == 7, "são 7 colunas, nem mais nem menos")

linha = [c.value for c in ws[2]]
checar(linha[2] == "Recepcionista", f"o CARGO está na coluna certa ({linha[2]!r})")
checar(linha[3] == nome_posto, f"o POSTO idem ({linha[3]!r})")
checar([linha[4], linha[5], linha[6]] == ["42", "M", "38"],
       f"as três MEDIDAS saem na ordem calça/camisa/calçado ({linha[4:7]})")

# --------------------------------------------------------------------------
print("\n2. o CPF sai MASCARADO, e vem da FICHA")
# --------------------------------------------------------------------------
# ⚠️ Mutações 1 e 4.
#
# Mascarado porque quem lê é uma pessoa, e `12345678909` num .xlsx ainda vira
# número (o Excel come o zero à esquerda). Da ficha porque é o CPF que a pessoa
# digitou e foi conferido contra o documento — o do convite é o que o RH
# digitou de memória ao convidar.
checar(linha[1] == "123.456.789-09",
       f"CPF com máscara, não cru ({linha[1]!r})")

cid2, _ = criar_pessoa(cpf_ficha="98765432100", cpf_convite="11111111111")
with SessionLocal() as db:
    cand2 = db.get(Candidato, uuid.UUID(cid2))
    l2 = linha_uniforme(db, cand2)
checar(l2["CPF"] == "987.654.321-00",
       f"com os dois preenchidos, vence o da FICHA (veio {l2['CPF']!r})")

cid3, _ = criar_pessoa(cpf_convite="11144477735")
with SessionLocal() as db:
    cand3 = db.get(Candidato, uuid.UUID(cid3))
    l3 = linha_uniforme(db, cand3)
checar(l3["CPF"] == "111.444.777-35",
       f"sem ficha, cai no do convite em vez de vir vazio ({l3['CPF']!r})")

cid4, _ = criar_pessoa(com_medidas=False)
with SessionLocal() as db:
    cand4 = db.get(Candidato, uuid.UUID(cid4))
    l4 = linha_uniforme(db, cand4)
checar(l4["CPF"] == "" and l4["Calça"] == "",
       "sem CPF e sem medidas, célula vazia é STRING vazia — `None` viraria a "
       "palavra 'None' na cara de quem abre")

# --------------------------------------------------------------------------
print("\n3. só o que serve para o UNIFORME")
# --------------------------------------------------------------------------
# Banco, PIX, endereço e salário estão a um `getattr` de distância. Anexo
# CIRCULA: o que não é necessário para a tarefa não deve viajar junto.
todo_texto = " ".join(str(c.value) for row in ws.iter_rows() for c in row)
for proibido in ("Banco do Brasil", "pix", "12345678909"):
    checar(proibido.lower() not in todo_texto.lower(),
           f"{proibido!r} NÃO aparece na planilha (minimização)")
checar(set(COLUNAS) == {"Nome", "CPF", "Cargo", "Posto", "Calça", "Camisa", "Calçado"},
       "o conjunto de colunas é exatamente o combinado — coluna nova aqui é "
       "dado pessoal a mais circulando por e-mail")

# --------------------------------------------------------------------------
print("\n4. falha no anexo NÃO segura o aviso")
# --------------------------------------------------------------------------
# ⚠️ Mutação 3: tirar o try/except -> a exceção sobe e o candidato não
# consegue concluir o envio. Perder o aviso (ou pior, travar a conclusão da
# admissão) por causa de um .xlsx seria trocar um problema pequeno por um maior
# — é a mesma regra do `avisar()`, que nunca levanta.
import pathlib  # noqa: E402
import re  # noqa: E402

fonte = (pathlib.Path(__file__).resolve().parents[1]
         / "app" / "api" / "documentos.py").read_text(encoding="utf-8")
trecho = fonte.split("uniforme_pendente")[0][-1600:]
checar("try:" in trecho and "except Exception" in trecho,
       "o monte do anexo está protegido por try/except no `concluir_envio`")
checar(re.search(r"anexos\s*=\s*None", trecho) is not None,
       "e o `anexos` nasce None — se a planilha falhar, o aviso sai sem ela")

# --------------------------------------------------------------------------
print("\n5. o nome do arquivo e o MIME")
# --------------------------------------------------------------------------
with SessionLocal() as db:
    cand = db.get(Candidato, uuid.UUID(cid))
    nome = nome_do_arquivo(cand)
checar(nome.startswith("uniforme-") and nome.endswith(".xlsx"),
       f"nome que se entende na caixa de entrada ({nome!r})")
checar("/" not in nome and "\\" not in nome and ".." not in nome,
       "passa pelo `slug()` — o nome é texto do usuário e vira NOME DE ARQUIVO")
tipo = _tipo_do_anexo(nome)
checar(tipo == ("application",
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
       f"o MIME é o do Excel, não `application/pdf` chumbado (veio {tipo}) — "
       f"o defeito que a v2.41 corrigiu no SMTP e a v2.68 no Graph")

# ⚠️ O MIME não pode depender de ONDE o código roda (v2.81, achado pelo CI):
# `mimetypes.guess_type` lê a tabela do SISTEMA (no Linux, /etc/mime.types), e
# a imagem do container não conhece `.xlsx`, `.ics` nem `.docx`. O teste passava
# no Windows e reprovava no CI — com o anexo saindo como `octet-stream`, que no
# caso do `.ics` significa convite SEM o "adicionar à agenda".
# Por isso o `_tipo_do_anexo` tem mapa EXPLÍCITO para o que o sistema não
# garante; estas asserções cobram que ele continue lá.
for ext, esperado in (
    (".xlsx", ("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
    (".ics", ("text", "calendar")),
    (".docx", ("application",
               "vnd.openxmlformats-officedocument.wordprocessingml.document")),
    (".md", ("text", "markdown")),
):
    veio = _tipo_do_anexo(f"arquivo{ext}")
    checar(veio == esperado,
           f"`{ext}` tem MIME próprio em QUALQUER sistema (veio {veio})")

# --------------------------------------------------------------------------
print("\n6. o TEXTO do e-mail diz que há anexo")
# --------------------------------------------------------------------------
# Até a v2.80 o corpo afirmava o contrário ("não vai por e-mail"). Instrução na
# interface que o sistema não cumpre é a armadilha da v2.74 — aqui, ao
# contrário: o sistema passou a cumprir e o texto tinha que acompanhar.
from app.services.email_templates import CATALOGO  # noqa: E402

modelo = next(m for m in CATALOGO if m.chave == "aviso_uniforme")
corpo = modelo.corpo.lower()
checar("anexo" in corpo,
       "o corpo menciona o ANEXO — sem isso o responsável não procuraria")
checar("não vai por e-mail" not in corpo,
       "e NÃO diz mais que os dados não vão por e-mail (era verdade até a v2.80)")


print()
if falhas:
    print(f"test_uniforme_planilha: {len(falhas)} FALHA(S)")
    for f_ in falhas:
        print(f"  - {f_}")
    raise SystemExit(1)
print("test_uniforme_planilha: OK")
