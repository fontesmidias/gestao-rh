"""Export de colaboradores para o DEXION — layout de conversão de trabalhadores.

Modelo oficial: `docs/orientações importacao dexion/PLANILHA MODELO CONVERSÃO
TRABALHADORES - dexion.xlsx`. O Bruno resumiu a exigência numa frase: *"siga
fielmente ao modelo da planilha, pois o dexion é mto enjoado"*.

NÃO reusa `montar_workbook_tirvu`, e isso é deliberado — os dois layouts se
parecem só de longe:

  ┌──────────────────┬───────────────────────┬──────────────────────────────┐
  │                  │ Tirvu                 │ Dexion                       │
  ├──────────────────┼───────────────────────┼──────────────────────────────┤
  │ colunas          │ 28                    │ 97 (A→CS)                    │
  │ aba              │ `Plan1`               │ `Sheet1`                     │
  │ cabeçalho        │ 1 linha               │ 4 linhas; dados na 5ª        │
  │ autoFilter       │ RECUSA                │ tem, em `A4:BV1039`          │
  │ datas            │ texto `dd/mm/aaaa`    │ SERIAL do Excel…             │
  │                  │                       │ …exceto INÍCIO (ESCALA)      │
  │ códigos          │ IDs da base do Tirvu  │ tabelas do eSocial           │
  └──────────────────┴───────────────────────┴──────────────────────────────┘

Copiar o gerador do Tirvu produziria um arquivo que PARECE certo e que o Dexion
recusa — ou pior, aceita com as datas erradas em mil e duzentos dias, porque
serial e texto são as duas coisas que um parser lê sem reclamar.

## A regra dos valores assumidos (decisão de 2026-08-02)

**Chumba-se o que é da EMPRESA; nunca o que é da PESSOA.**

O que é invariante da empregadora — país, moeda, regime da empresa, tipo de
declaração — vai com valor fixo, pelo mesmo raciocínio que já rege o
`EMPRESA_TIRVU_ID = "1"`: o grupo opera com uma empregadora só.

O que varia por pessoa — categoria do trabalhador, CBO, sindicato, conta
bancária, grau de instrução — **vira PENDÊNCIA anunciada**, nunca padrão
silencioso. Essa distinção não é preciosismo: um código eSocial errado não dá
erro na importação. Entra limpo e sai errado na declaração ao governo, meses
depois, sem nada acusar. É a mesma assinatura de defeito do "Registra Ponto"
que mordeu na v1.82 — célula em branco que o importador aceitou calado e o
colaborador nasceu lá sem a marcação.

## O que o sistema ainda NÃO coleta

Agência, conta e tipo de conta bancária (o cadastro tem só `banco` em texto
livre e a chave PIX), município IBGE e CBO por pessoa. Decisão do Bruno em
2026-08-02: **exportar vazio por ora** e tratar como pendência anunciada — o
alinhamento sobre quais colunas o Dexion realmente exige é conversa com eles.
Inventar dado bancário seria muito pior que entregar a coluna em branco.
"""

import io
from datetime import date, datetime
from typing import TYPE_CHECKING

if TYPE_CHECKING:                       # só para tipagem; não roda em runtime
    from sqlalchemy.orm import Session

    from app.models.candidato import Candidato

# NADA de import de modelo no topo. `app.models` puxa `app.core.config`, que
# exige `pydantic_settings` — e isso tornaria impossível testar a FORMA da
# planilha (as 97 colunas, o serial das datas, o cabeçalho contra o modelo
# oficial) sem instalar meia aplicação. O CI pagou esse preço uma vez.
#
# Os imports de modelo ficam dentro de `linha_dexion`, que é a única função que
# realmente toca o banco. `montar_workbook_dexion`, `pendencias_linha` e os
# formatadores são puros e importáveis com openpyxl e stdlib apenas.

# ---------------------------------------------------------------------------
# As 97 colunas, na ordem EXATA do modelo (A→CS). É a única fonte de verdade da
# ordem: o gerador itera sobre esta lista, nunca sobre as chaves do dict — a
# união das chaves sairia em ordem de inserção e o Dexion lê por POSIÇÃO.
#
# Os nomes repetem-se entre grupos ("CATEGORIA" aparece em Documentos, Admissão
# e Sefip; "UF" em Nascimento, Documentos e Endereço; "TIPO DE JORNADA" duas
# vezes). Por isso a chave do dict é a LETRA DA COLUNA, não o rótulo — com
# rótulo, três colunas diferentes colidiriam numa só.
# ---------------------------------------------------------------------------
COLUNAS_DEXION: list[tuple[str, str, str]] = [
    # (letra, grupo da linha 3, rótulo da linha 4)
    ("A",  "Cabeçalho", "MATRÍCULA"),
    ("B",  "Cabeçalho", "NOME"),
    ("C",  "Cabeçalho", "CPF"),
    ("D",  "Cabeçalho", "MATRÍCULA ESOCIAL"),
    ("E",  "Dados  cadastrais> Gerais", "SEXO"),
    ("F",  "Dados  cadastrais> Gerais", "ETNIA/RAÇA"),
    ("G",  "Dados  cadastrais> Gerais", "TIPO SANGUÍNEO"),
    ("H",  "Dados  cadastrais> Gerais", "GRAU DE INSTRUÇÃO"),
    ("I",  "Dados  cadastrais> Gerais", "ESTADO CIVIL"),
    ("J",  "Dados  cadastrais> Gerais", "NACIONALIDADE"),
    ("K",  "Dados cadastrais> Gerais> Nascimento", "NASCIMENTO"),
    ("L",  "Dados cadastrais> Gerais> Nascimento", "NATURALIDADE"),
    ("M",  "Dados cadastrais> Gerais> Nascimento", "UF"),
    ("N",  "Dados cadastrais> Gerais> Nascimento", "NOME DO PAI"),
    ("O",  "Dados cadastrais> Gerais> Nascimento", "NOME DA MÃE"),
    ("P",  "Dados cadastrais> Documentos", "TIPO DE IDENTIDADE"),
    ("Q",  "Dados cadastrais> Documentos", "NÚMERO (IDENTIDADE)"),
    ("R",  "Dados cadastrais> Documentos", "ÓRGÃO EMISSOR"),
    ("S",  "Dados cadastrais> Documentos", "ÓRGÃO EMISSOR (uf)"),
    ("T",  "Dados cadastrais> Documentos", "DATA EXPED. (IDENTIDADE)"),
    ("U",  "Dados cadastrais> Documentos", "PISPASEP"),
    ("V",  "Dados cadastrais> Documentos", "CNH"),
    ("W",  "Dados cadastrais> Documentos", "CATEGORIA"),
    ("X",  "Dados cadastrais> Documentos", "VALIDADE (CNH)"),
    ("Y",  "Dados cadastrais> Documentos", "ÓRGÃO EMISSOR (CNH)"),
    ("Z",  "Dados cadastrais> Documentos", "TÍTULO ELEITORAL"),
    ("AA", "Dados cadastrais> Documentos", "ZONA"),
    ("AB", "Dados cadastrais> Documentos", "SEÇÃO"),
    ("AC", "Dados cadastrais> Documentos", "CTPS"),
    ("AD", "Dados cadastrais> Documentos", "SÉRIE"),
    ("AE", "Dados cadastrais> Documentos", "UF (CTPS)"),
    ("AF", "Dados cadastrais> Documentos", "DATA EXPED. (CTPS)"),
    ("AG", "Dados cadastrais> Documentos", "RESERVISTA"),
    ("AH", "Dados cadastrais> Documentos", "CATEGORIA (RESERVISTA)"),
    ("AI", "Dados cadastrais> Documentos", "TIPO DA CONTA BANCÁRIA"),
    ("AJ", "Dados cadastrais> Documentos", "BANCO"),
    ("AK", "Dados cadastrais> Documentos", "AGÊNCIA"),
    ("AL", "Dados cadastrais> Documentos", "CONTA"),
    ("AM", "Dados cadastrais> Endereço", "PAIS"),
    ("AN", "Dados cadastrais> Endereço", "ENDEREÇO"),
    ("AO", "Dados cadastrais> Endereço", "NÚMERO"),
    ("AP", "Dados cadastrais> Endereço", "COMPLEMENTO"),
    ("AQ", "Dados cadastrais> Endereço", "BAIRRO"),
    ("AR", "Dados cadastrais> Endereço", "CIDADE"),
    ("AS", "Dados cadastrais> Endereço", "UF"),
    ("AT", "Dados cadastrais> Endereço", "CEP"),
    ("AU", "Dados cadastrais> Endereço", "MUNICÍPIO"),
    ("AV", "Dados cadastrais> Endereço", "TELEFONE PRINCIPAL"),
    ("AW", "Dados cadastrais> Endereço", "TELEFONE ALTERNATIVO"),
    ("AX", "Dados cadastrais> Endereço", "EMAIL"),
    ("AY", "Dados contratuais> Admissão", "ADMISSÃO"),
    ("AZ", "Dados contratuais> Admissão", "CATEGORIA"),
    ("BA", "Dados contratuais> Admissão", "OPÇÃO PELO FGTS"),
    ("BB", "Dados contratuais> Admissão", "DATA OPÇÃO"),
    ("BC", "Dados contratuais> Admissão", "NATUREZA DA ATIVIDADE"),
    ("BD", "Dados contratuais> Admissão", "TIPO DE ADMISSÃO"),
    ("BE", "Dados contratuais> Admissão", "TIPO DE CONTRATO"),
    ("BF", "Dados contratuais> Admissão", "CONTRATO DE EXPERIÊNCIA"),
    ("BG", "Dados contratuais> Admissão", "DIAS EXPERIÊNCIA"),
    ("BH", "Dados contratuais> Admissão", "DIAS PRORROGAÇÃO"),
    ("BI", "Dados contratuais> Admissão> Empregado", "TEMPO PARCIAL"),
    ("BJ", "Dados contratuais> Admissão> Empregado", "MEIO DE ADMISSÃO"),
    ("BK", "Dados contratuais> Admissão> Empregado", "REGIME TRABALHISTA"),
    ("BL", "Dados contratuais> Admissão> Empregado", "REGIME PREVIDENCIÁRIO"),
    ("BM", "Dados contratuais> Outros", "UNIDADE DO SALÁRIO FIXO"),
    ("BN", "Dados contratuais> Outros", "CARGO"),
    ("BO", "Dados contratuais> Outros", "CBO"),
    ("BP", "Dados contratuais> Outros", "SALÁRIO FIXO"),
    ("BQ", "Dados contratuais> Outros", "HORAS SEMANAIS"),
    ("BR", "Dados contratuais> Outros", "HORAS MENSAIS"),
    ("BS", "Dados contratuais> Outros> Sindicato", "SINDICATO"),
    ("BT", "Dados contratuais> Outros> Sindicato", "MÊS DATA-BASE"),
    ("BU", "Eventos trabalhistas> Diversos> Lotações", "DATA"),
    ("BV", "Eventos trabalhistas> Diversos> Lotações", "LOTAÇÃO"),
    ("BW", "Eventos trabalhistas> Diversos> Lotações", "TIPO DE JORNADA"),
    ("BX", "Eventos trabalhistas> Diversos> Lotações", "DIAS DE TRABALHO"),
    ("BY", "Eventos trabalhistas> Diversos> Lotações", "DIAS DE DESCANSO"),
    ("BZ", "Eventos trabalhistas> Diversos> Lotações", "INÍCIO (ESCALA)"),
    ("CA", "Eventos trabalhistas> Diversos> Lotações", "TIPO DE JORNADA (eSocial)"),
    ("CB", "Eventos trabalhistas> Diversos> Lotações", "REGIME DE JORNADA"),
    ("CC", "Eventos trabalhistas> Diversos> Lotações", "GRAU DE EXPOSIÇÃO A AGENTES NOCIVOS"),
    ("CD", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "TIPO DE JORNADA"),
    ("CE", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "JORNADA"),
    ("CF", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "ITINERÁRIO"),
    ("CG", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "QTDE DIÁRIA (ITINERÁRIO)"),
    ("CH", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "BENEFÍCIO VALE REFEIÇÃO"),
    ("CI", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "QTDE DIÁRIA (VALE REFEIÇÃO)"),
    ("CJ", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "BENEFÍCIO VALE ALIMENTAÇÃO"),
    ("CK", "Eventos trabalhistas> Diversos> Lotações> Jornadas de trabalho", "QTDE DIÁRIA (VALE ALIMENTAÇÃO)"),
    ("CL", "Compatibilidade> Sefip", "CATEGORIA"),
    ("CM", "Compatibilidade> Sefip", "OCORRÊNCIA"),
    ("CN", "Compatibilidade> Caged", "MOVIMENTAÇÃO DA ADMISSÃO"),
    ("CO", "Compatibilidade> Caged", "TIPO DE DECLARAÇÃO DA ADMISSÃO"),
    ("CP", "Compatibilidade> RAIS", "NACIONALIDADE"),
    ("CQ", "Compatibilidade> RAIS", "TIPO DE ADMISSÃO"),
    ("CR", "Compatibilidade> RAIS", "VÍNCULO"),
    ("CS", "Compatibilidade> RAIS", "TIPO DE SALÁRIO"),
]

ABA_DEXION = "Sheet1"
TITULO_L1 = "EMPRESA XPTO"          # o modelo traz o nome da empresa aqui
SUBTITULO_L2 = "DADOS CADASTRAIS FUNCIONARIOS IMPORTACAO"
LINHA_GRUPOS = 3
LINHA_CABECALHO = 4
PRIMEIRA_LINHA_DADOS = 5

# ---------------------------------------------------------------------------
# Valores fixos: SÓ o que é invariante da EMPRESA (ver a regra no topo).
# ---------------------------------------------------------------------------
PAIS_BRASIL = "105"          # AM — tabela 6 do eSocial
NACIONALIDADE_BRASILEIRA = "1"   # J
RAIS_NACIONALIDADE_BR = "10"     # CP — a RAIS usa codificação própria
OPCAO_FGTS = "Sim"           # BA — regime CLT: todo empregado é optante
REGIME_TRABALHISTA_CLT = "1"     # BK
REGIME_PREVIDENCIARIO_RGPS = "1"  # BL
CAGED_MOVIMENTACAO_ADMISSAO = "20"   # CN — admissão de trabalhador
CAGED_TIPO_DECLARACAO = "1"          # CO

# ---------------------------------------------------------------------------
# De-para dos enums do sistema para os códigos do eSocial.
#
# Estes NÃO são "valores chumbados": são tradução de um dado que a pessoa
# informou. O que não tiver correspondência sai VAZIO e vira pendência — nunca
# um código chutado, porque código errado entra calado na declaração.
# ---------------------------------------------------------------------------

# Tabela 10 do eSocial (grau de instrução). O enum do sistema é mais grosso que
# a tabela oficial; mapeia-se para o código mais próximo que não afirma demais.
GRAU_INSTRUCAO = {
    "fund_incompleto": "02",
    "fund_completo": "04",
    "medio_incompleto": "05",
    "medio_completo": "06",
    "sup_incompleto": "07",
    "sup_completo": "08",
    "pos_graduacao": "09",
}

# Tabela 7 do eSocial (raça/cor). `nao_informar` não existe no enum de cor do
# sistema; ausência sai vazia.
ETNIA_RACA = {
    "branca": "1", "preta": "2", "parda": "3", "amarela": "4", "indigena": "5",
}

# Estado civil: codificação do próprio Dexion (o eSocial não tem tabela para
# isto). Segue o exemplo do modelo, em que solteiro = 1.
ESTADO_CIVIL = {
    "solteiro": "1", "casado": "2", "divorciado": "3", "separado": "4",
    "viuvo": "5", "uniao_estavel": "6",
}


def _serial_excel(v) -> int | None:
    """Data → serial do Excel (dias desde 30/12/1899).

    O Dexion espera as datas como NÚMERO com formato `dd/mm/yyyy` — é assim que
    o modelo traz `K=32874` e `AY=46082`. Mandar texto "12/10/1998" numa célula
    de data faz o Excel tratar como string: alinha à esquerda, não ordena, e o
    importador lê outra coisa (ou nada).

    A exceção é `BZ` (INÍCIO DA ESCALA), que no modelo é TEXTO `01/02/2026` —
    ver `_linha_dexion`. Duas convenções na mesma planilha; seguir o modelo é
    mais seguro que uniformizar.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, str):
        texto = v.strip()
        try:
            if "-" in texto:                       # ISO aaaa-mm-dd
                a, m, d = (int(x) for x in texto.split("-"))
            else:                                  # BR dd/mm/aaaa
                d, m, a = (int(x) for x in texto.split("/"))
            v = date(a, m, d)
        except (ValueError, TypeError):
            return None
    if not isinstance(v, date):
        return None
    # 30/12/1899 é a época do Excel (o "bug" do ano 1900 já embutido).
    return (v - date(1899, 12, 30)).days


def _salario_texto(bruto: str | None) -> str:
    """`R$ 1.500,00` → `1500,00`, como o modelo (`BP=2000,00`, célula TEXTO).

    Devolve string vazia quando não dá para interpretar — nunca zero. Salário
    zero numa importação de folha seria aceito calado e sairia no contracheque.
    """
    if not bruto:
        return ""
    texto = str(bruto).strip().replace("R$", "").replace(" ", "")
    if not texto:
        return ""
    # pt-BR: ponto é milhar, vírgula é decimal.
    normal = texto.replace(".", "").replace(",", ".")
    try:
        return f"{float(normal):.2f}".replace(".", ",")
    except ValueError:
        return ""


def linha_dexion(db: "Session", c: "Candidato", gerar_matricula: bool = False) -> dict:
    """Uma linha do layout, com a LETRA da coluna como chave.

    Chave por letra, e não por rótulo, porque o layout REPETE nomes: "CATEGORIA"
    aparece três vezes (documentos, admissão, Sefip), "UF" três, "TIPO DE
    JORNADA" duas. Com rótulo, uma sobrescreveria a outra em silêncio e o
    arquivo sairia com a coluna errada preenchida.

    `gerar_matricula=True` só no EXPORT (grava); a pré-checagem passa False —
    consulta não muta dados. Mesma regra do `linha_tirvu`.
    """
    # Imports AQUI, não no topo: é o que mantém o resto do módulo testável sem
    # `pydantic_settings` e o resto da aplicação (ver a nota no cabeçalho).
    from app.models.candidato import Jornada, PostoServico
    from app.models.ficha import (DadosPessoais, DocumentosIdentificacao,
                                  Endereco)
    from app.services.export_tirvu import _so_digitos, garantir_matricula

    p = db.get(DadosPessoais, c.id)
    e = db.get(Endereco, c.id)
    d = db.get(DocumentosIdentificacao, c.id)
    posto = db.get(PostoServico, c.posto_servico_id) if c.posto_servico_id else None
    jornada = db.get(Jornada, c.jornada_id) if c.jornada_id else None

    cpf = _so_digitos((d.cpf if d and d.cpf else c.cpf) or "")
    # Fallback duplo: `DadosPessoais.data_nascimento` é Date e
    # `Candidato.data_nascimento` é String(10) — os dois formatos coexistem.
    nascimento = (p.data_nascimento if p and p.data_nascimento
                  else c.data_nascimento)

    sexo = ""
    if p and p.sexo:
        sexo = "M" if p.sexo.value == "masculino" else "F"

    # Endereço nos DOIS formatos (v2.37): a coleta atual grava os campos
    # separados e deixa o legado nulo; a antiga tem só a string única.
    logradouro, numero, complemento = "", "", ""
    if e:
        if e.logradouro:
            logradouro, numero, complemento = (
                e.logradouro, e.numero or "", e.complemento or "")
        else:
            logradouro = e.logradouro_numero_complemento or ""

    matricula = garantir_matricula(db, c) if gerar_matricula else (c.matricula or "")

    def enum_val(campo) -> str:
        """Valor do enum como string, ou "" — enums podem vir None."""
        return campo.value if campo is not None else ""

    return {
        # --- Cabeçalho ---
        "A": matricula,
        "B": c.nome_completo or "",
        "C": cpf,                      # sem máscara (modelo: 95315637050)
        # MATRÍCULA ESOCIAL no modelo é `0001-000001` (empresa-matrícula).
        # Sem empresa cadastrada no Dexion, não se inventa o prefixo.
        "D": "",
        # --- Gerais ---
        "E": sexo,
        "F": ETNIA_RACA.get(enum_val(p.cor_raca) if p else "", ""),
        "G": _tipo_sanguineo(db, c),
        "H": GRAU_INSTRUCAO.get(enum_val(p.escolaridade) if p else "", ""),
        "I": ESTADO_CIVIL.get(enum_val(p.estado_civil) if p else "", ""),
        # Nacionalidade: só se declarada brasileira. Estrangeiro tem código
        # próprio por país e o sistema não guarda qual — sai vazio (pendência),
        # nunca "brasileiro" por omissão.
        "J": (NACIONALIDADE_BRASILEIRA
              if (p and p.nacionalidade and p.nacionalidade.value == "brasileira") else ""),
        "K": _serial_excel(nascimento),
        "L": (p.naturalidade_cidade if p else "") or "",
        "M": (p.naturalidade_uf if p else "") or "",
        "N": (p.nome_pai if p else "") or "",
        "O": (p.nome_mae if p else "") or "",
        # --- Documentos ---
        "P": "RG" if (d and d.rg_numero) else "",
        "Q": (d.rg_numero if d else "") or "",
        "R": (d.rg_orgao_emissor if d else "") or "",
        "S": "",                       # UF do órgão emissor: não coletada
        "T": _serial_excel(d.rg_data_expedicao if d else None),
        "U": _so_digitos(d.pis_nis_pasep if d else ""),
        "V": (d.cnh_numero if d else "") or "",
        "W": (d.cnh_categoria if d else "") or "",
        "X": _serial_excel(d.cnh_validade if d else None),
        "Y": (d.cnh_orgao_emissor if d else "") or "",
        "Z": (d.titulo_eleitor_numero if d else "") or "",
        "AA": (d.titulo_eleitor_zona if d else "") or "",
        "AB": (d.titulo_eleitor_secao if d else "") or "",
        "AC": (d.ctps_numero if d else "") or "",
        "AD": (d.ctps_serie if d else "") or "",
        "AE": "",                      # UF da CTPS: não coletada
        "AF": None,                    # data de expedição da CTPS: não coletada
        "AG": (d.militar_numero if d else "") or "",
        "AH": (d.militar_categoria if d else "") or "",
        # Conta bancária: o sistema guarda só `banco` (texto livre) e a chave
        # PIX. Agência, conta e tipo NÃO são coletados — decisão de 2026-08-02
        # foi exportar vazio e tratar como pendência anunciada, em vez de
        # inventar dado que decide para onde vai o salário.
        "AI": "",
        "AJ": "",
        "AK": "",
        "AL": "",
        # --- Endereço ---
        "AM": PAIS_BRASIL,
        "AN": logradouro,
        "AO": numero,
        "AP": complemento,
        "AQ": (e.bairro if e else "") or "",
        "AR": (e.cidade if e else "") or "",
        "AS": (e.uf if e else "") or "",
        "AT": _so_digitos(e.cep if e else ""),   # modelo: 72130000, sem hífen
        "AU": "",                      # código IBGE do município: não coletado
        "AV": _so_digitos(c.celular_whatsapp),
        "AW": "",
        "AX": c.email or "",
        # --- Admissão ---
        "AY": _serial_excel(c.data_admissao),
        "AZ": "",                      # categoria do trabalhador: por pessoa
        "BA": OPCAO_FGTS,
        "BB": _serial_excel(c.data_admissao),    # opção pelo FGTS na admissão
        "BC": "",                      # natureza da atividade: por posto
        "BD": "",                      # tipo de admissão: por pessoa
        "BE": "",                      # tipo de contrato: por pessoa
        "BF": "",
        "BG": "",
        "BH": "",
        "BI": "",
        "BJ": "",
        "BK": REGIME_TRABALHISTA_CLT,
        "BL": REGIME_PREVIDENCIARIO_RGPS,
        # --- Outros ---
        "BM": "",                      # unidade do salário fixo: por contrato
        "BN": c.cargo_funcao or "",
        "BO": _cbo_do_cargo(db, c.cargo_funcao),
        "BP": _salario_texto(c.salario_base),
        "BQ": "",
        "BR": "",
        "BS": "",                      # sindicato: por posto/categoria
        "BT": "",
        # --- Lotações ---
        "BU": _serial_excel(c.data_admissao),
        "BV": (posto.tirvu_id if posto else "") or "",
        "BW": "",
        "BX": "",
        "BY": "",
        # ATENÇÃO: TEXTO dd/mm/aaaa, não serial — é assim no modelo (BZ do
        # exemplo é `01/02/2026`, célula de texto), ao contrário de todas as
        # outras datas da planilha.
        "BZ": _data_br_texto(c.data_admissao),
        "CA": "",
        "CB": "",
        "CC": "",
        # --- Jornadas ---
        "CD": "",
        "CE": (jornada.descricao if jornada else "") or "",
        "CF": "",
        "CG": "",
        "CH": "",
        "CI": "",
        "CJ": "",
        "CK": "",
        # --- Compatibilidade ---
        "CL": "",                      # categoria Sefip: por pessoa
        "CM": "",
        "CN": CAGED_MOVIMENTACAO_ADMISSAO,
        "CO": CAGED_TIPO_DECLARACAO,
        "CP": (RAIS_NACIONALIDADE_BR
               if (p and p.nacionalidade and p.nacionalidade.value == "brasileira") else ""),
        "CQ": "",
        "CR": "",
        "CS": "",
    }


def _tipo_sanguineo(db: "Session", c: "Candidato") -> str:
    """Tipo sanguíneo (coluna G) — mora na ficha de emergência.

    Dado de saúde (LGPD art. 11): vai para o Dexion porque o sistema de folha é
    onde a ficha de emergência do trabalhador é consultada, mas note que é o
    ÚNICO campo sensível deste layout — não acrescentar outros sem conferir a
    base legal.
    """
    from app.models.ficha import FichaEmergencia
    ficha = db.get(FichaEmergencia, c.id)
    return (ficha.tipo_sanguineo if ficha else "") or ""


def _cbo_do_cargo(db: "Session", cargo: str | None) -> str:
    """CBO a partir do de-para de cargos do Tirvu, que já guarda o código.

    Reusa `CargoTirvu.cbo` em vez de criar um cadastro paralelo: o RH já
    preenche esse de-para, e duas tabelas de CBO divergiriam na primeira
    correção feita só num lado. Sem correspondência, devolve "" — o CBO define
    a ocupação declarada ao governo e não se chuta.
    """
    if not cargo:
        return ""
    from app.models.candidato import CargoTirvu
    from app.services.export_tirvu import normalizar_cargo
    from sqlalchemy import select
    reg = db.scalar(select(CargoTirvu)
                    .where(CargoTirvu.cargo_normalizado == normalizar_cargo(cargo)))
    return (reg.cbo if reg and reg.cbo else "") or ""


def _data_br_texto(v) -> str:
    """`dd/mm/aaaa` como TEXTO — só para a coluna BZ (início da escala)."""
    s = _serial_excel(v)
    if s is None:
        return ""
    from datetime import timedelta
    d = date(1899, 12, 30) + timedelta(days=s)
    return d.strftime("%d/%m/%Y")


def montar_workbook_dexion(linhas: list[dict],
                           nome_empresa: str = TITULO_L1) -> bytes:
    """Gera o .xlsx FIEL ao modelo do Dexion.

    O que o modelo exige, e que difere do gerador do Tirvu em todos os pontos:

    * aba `Sheet1`;
    * linha 1 com o nome da empresa, linha 2 com o subtítulo;
    * linha 3 com os GRUPOS e linha 4 com os rótulos das colunas;
    * dados a partir da linha 5;
    * `autoFilter` em `A4:BV<fim>` — o Tirvu RECUSA autofiltro, o Dexion tem;
    * datas como número com formato `dd/mm/yyyy`; texto para CPF, PIS, CEP e
      salário (senão o Excel come o zero à esquerda do CEP e da matrícula).

    Célula vazia é PULADA — o openpyxl geraria `<c t="inlineStr"></c>`
    malformado, que parsers rígidos recusam (lição do Tirvu, v1.79).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = ABA_DEXION

    ws.cell(row=1, column=1, value=nome_empresa).font = Font(bold=True)
    ws.cell(row=2, column=1, value=SUBTITULO_L2).font = Font(bold=True)
    for j, (_letra, grupo, rotulo) in enumerate(COLUNAS_DEXION, start=1):
        ws.cell(row=LINHA_GRUPOS, column=j, value=grupo)
        ws.cell(row=LINHA_CABECALHO, column=j, value=rotulo).font = Font(bold=True)

    # Colunas cujo valor é DATA (serial + formato). BZ fica de fora de
    # propósito: no modelo ela é texto.
    COLUNAS_DATA = {"K", "T", "X", "AF", "AY", "BB", "BU"}
    # Texto puro: preserva zero à esquerda (CEP 72130000, matrícula 000001) e
    # impede o Excel de "ajudar" convertendo CPF em notação científica.
    COLUNAS_TEXTO = {"A", "C", "D", "U", "AT", "BP", "AV", "AW", "Q", "AC"}

    for i, linha in enumerate(linhas, start=PRIMEIRA_LINHA_DADOS):
        for j, (letra, _grupo, _rotulo) in enumerate(COLUNAS_DEXION, start=1):
            v = linha.get(letra, "")
            if v is None or v == "":
                continue
            cel = ws.cell(row=i, column=j, value=v)
            if letra in COLUNAS_DATA:
                cel.number_format = "dd/mm/yyyy"
            elif letra in COLUNAS_TEXTO:
                cel.number_format = "@"

    # O modelo tem autoFilter em A4:BV — sobre a linha de CABEÇALHO (4), não a
    # do título. Vai até BV mesmo (não CS): é o que está no arquivo oficial.
    ultima = max(PRIMEIRA_LINHA_DADOS - 1, len(linhas) + PRIMEIRA_LINHA_DADOS - 1)
    ws.auto_filter.ref = f"A{LINHA_CABECALHO}:BV{ultima}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Pendências: o que falta para o Dexion aceitar a linha.
#
# Rótulo AMIGÁVEL, não o nome da coluna: o RH lê "Conta bancária", não "AK".
# Mesma mecânica do `_ROTULO_PENDENCIA` do Tirvu.
# ---------------------------------------------------------------------------
_OBRIGATORIAS = [
    ("B", "Nome completo"),
    ("C", "CPF"),
    ("K", "Data de nascimento"),
    ("U", "PIS/PASEP"),
    ("AY", "Data de admissão"),
    ("BN", "Cargo"),
    ("BO", "CBO do cargo (cadastre em Configurações → Cargos)"),
    ("BP", "Salário"),
    ("AN", "Endereço"),
    ("AT", "CEP"),
    ("BV", "Lotação (ID do posto)"),
    ("H", "Grau de instrução"),
    ("O", "Nome da mãe"),
]


def pendencias_linha(linha: dict) -> list[str]:
    """Campos vazios que o Dexion precisa, em rótulo que o RH entende.

    Opera sobre a LINHA MONTADA, não sobre o `Candidato`: garante que o que se
    confere é exatamente o que vai sair na célula, e mantém a função pura e
    testável sem banco.
    """
    faltam = []
    for letra, rotulo in _OBRIGATORIAS:
        v = linha.get(letra)
        if v is None or (isinstance(v, str) and not v.strip()):
            faltam.append(rotulo)
    return faltam
