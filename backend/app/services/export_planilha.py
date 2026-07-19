"""Geração de planilhas Excel e utilitários de texto compartilhados entre a
página de Colaboradores e o menu de Arquivo.

`linha_completa` acha a ficha inteira de um colaborador; `montar_workbook`
recebe uma lista de DICTS já prontos (não `Candidato`) e devolve os bytes do
XLSX — assim o chamador injeta colunas extras (vínculo, contagens de arquivo)
sem duplicar a montagem. `slug` normaliza texto para uso seguro em nomes de
arquivo/pasta (path-safe)."""

import io
import re
import unicodedata

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.candidato import Candidato
from app.models.ficha import (ContatoEmergencia, DadosPessoais,
                              DadosProfissionaisBancarios, Dependente,
                              DocumentosIdentificacao, Endereco, FichaEmergencia,
                              ValeTransporte)

# Nomes reservados do Windows (o RH abre o ZIP no Windows).
_RESERVADOS_WINDOWS = {
    "CON", "PRN", "AUX", "NUL", *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def normalizar(txt) -> str:
    """Sem acento, minúsculo, espaços colapsados (base de casamento de texto)."""
    txt = unicodedata.normalize("NFKD", str(txt or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", txt).strip().lower()


def slug(txt, fallback: str = "sem-nome", maximo: int = 80) -> str:
    """Componente de caminho SEGURO: remove acentos, barras, pontos e '..' —
    nunca deixa escapar path traversal. Colapsa para o fallback se ficar vazio."""
    base = unicodedata.normalize("NFKD", str(txt or "")).encode("ascii", "ignore").decode()
    # troca tudo que não for alfanumérico/hífen/underscore por hífen; isso já
    # elimina '/', '\\', ':', '.', espaços e sequências '..'
    base = re.sub(r"[^A-Za-z0-9_-]+", "-", base)
    base = re.sub(r"-+", "-", base).strip("-")[:maximo].strip("-")
    if not base or base.upper() in _RESERVADOS_WINDOWS:
        return fallback
    return base


def _fmt(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, bool):
        return "Sim" if v else "Não"
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    if hasattr(v, "value"):
        return str(v.value).replace("_", " ")
    return str(v)


def linha_completa(db: Session, c: Candidato) -> dict:
    """Todas as respostas do candidato, achatadas em um dicionário ordenado."""
    p = db.get(DadosPessoais, c.id)
    e = db.get(Endereco, c.id)
    d = db.get(DocumentosIdentificacao, c.id)
    b = db.get(DadosProfissionaisBancarios, c.id)
    vt = db.get(ValeTransporte, c.id)
    fe = db.get(FichaEmergencia, c.id)
    deps = db.scalars(select(Dependente).where(Dependente.candidato_id == c.id)).all()
    contatos = db.scalars(select(ContatoEmergencia)
                          .where(ContatoEmergencia.candidato_id == c.id)).all()

    linha = {
        "Nome completo": c.nome_completo, "E-mail": c.email,
        "Celular/WhatsApp": c.celular_whatsapp, "Status": _fmt(c.status),
        "Convidado em": _fmt(c.criado_em), "Dossiê gerado em": _fmt(c.dossie_gerado_em),
    }
    if p:
        linha.update({
            "Nome social": _fmt(p.nome_social),
            "Filiação (mãe)": _fmt(p.nome_mae), "Filiação (pai)": _fmt(p.nome_pai),
            "Data de nascimento": _fmt(p.data_nascimento), "Sexo": _fmt(p.sexo),
            "Identidade de gênero": _fmt(p.identidade_genero), "Cor/raça": _fmt(p.cor_raca),
            "Nacionalidade": _fmt(p.nacionalidade),
            "Naturalidade (cidade)": _fmt(p.naturalidade_cidade),
            "Naturalidade (UF)": _fmt(p.naturalidade_uf),
            "Estado civil": _fmt(p.estado_civil), "Escolaridade": _fmt(p.escolaridade),
            "PCD": _fmt(p.pcd),
        })
    if e:
        linha.update({
            "CEP": _fmt(e.cep), "Endereço": _fmt(e.logradouro_numero_complemento),
            "Bairro": _fmt(e.bairro), "Cidade": _fmt(e.cidade), "UF": _fmt(e.uf),
        })
    if d:
        linha.update({
            "RG": _fmt(d.rg_numero), "RG órgão emissor": _fmt(d.rg_orgao_emissor),
            "RG expedição": _fmt(d.rg_data_expedicao), "CPF": _fmt(d.cpf),
            "PIS/NIS/PASEP": _fmt(d.pis_nis_pasep), "CNH": _fmt(d.cnh_numero),
            "CNH categoria": _fmt(d.cnh_categoria),
            "Título de eleitor": _fmt(d.titulo_eleitor_numero),
            "Título zona": _fmt(d.titulo_eleitor_zona),
            "Título seção": _fmt(d.titulo_eleitor_secao),
        })
    if b:
        linha.update({
            "Tam. calça": _fmt(b.tamanho_calca), "Tam. camisa": _fmt(b.tamanho_camisa),
            "Tam. calçado": _fmt(b.tamanho_calcado), "Banco": _fmt(b.banco),
            "PIX tipo": _fmt(b.pix_tipo), "PIX chave": _fmt(b.pix_chave),
        })
    if vt:
        linha.update({
            "VT optante": _fmt(vt.optante), "VT cartão": _fmt(vt.cartao_dftrans),
            "VT trajeto": _fmt(vt.trajeto_descricao),
        })
    if fe:
        linha.update({
            "Tipo sanguíneo": _fmt(fe.tipo_sanguineo),
            "Medicamento contínuo": _fmt(fe.usa_medicamento_continuo),
            "Medicamentos": _fmt(fe.medicamentos),
            "Condições médicas": _fmt(fe.condicoes_medicas),
            "Orientações de emergência": _fmt(fe.orientacao_emergencia),
        })
    linha["Dependentes"] = "; ".join(
        f"{dep.nome_completo} ({_fmt(dep.parentesco)}, nasc. {_fmt(dep.data_nascimento)}, "
        f"CPF {dep.cpf}{', deduz IRRF' if dep.deduz_irrf else ''})"
        for dep in deps
    )
    linha["Contatos de emergência"] = "; ".join(
        f"{ct.nome_completo} ({ct.parentesco}, {ct.telefone_celular})"
        for ct in contatos
    )
    return linha


def montar_workbook(linhas: list[dict], titulo: str = "Colaboradores") -> bytes:
    """Recebe DICTS já prontos e devolve os bytes do XLSX. As colunas são a
    UNIÃO de todas as chaves na ordem em que aparecem (linhas incompletas não
    escondem colunas das completas)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    colunas: list[str] = []
    for linha in linhas:
        for chave in linha:
            if chave not in colunas:
                colunas.append(chave)

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # limite do Excel para nome de aba
    verde = PatternFill("solid", fgColor="0FB257")
    for j, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=j, value=nome)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = verde
        cel.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = max(14, min(38, len(nome) + 6))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(len(colunas), 1))}1"
    for i, linha in enumerate(linhas, start=2):
        for j, nome in enumerate(colunas, start=1):
            ws.cell(row=i, column=j, value=linha.get(nome, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
