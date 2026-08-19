"""Planilha de importação de admissões do Tirvu ("Layout de Importação de
Admissões"): 28 colunas em ordem FIXA — o RH baixa daqui e sobe lá, sem
redigitar. O Tirvu deduplica por CPF do lado dele (quem já existe é ignorado)
e RECUSA linha sem CTPS e sem PIS.

Formatos conforme a exportação do próprio Tirvu: CPF com máscara, datas
dd/mm/aaaa. CTPS Digital derivada do CPF no formato que o Tirvu importa
(feedback 2026-07-24): número = os 7 primeiros dígitos do CPF, série = os 4
últimos (juntos reconstroem o CPF). NÃO é (CPF completo, "0000")."""

import io
import re
import unicodedata

from sqlalchemy.orm import Session

from app.models.candidato import Candidato, Empresa, Jornada, PostoServico
from app.models.ficha import (DadosPessoais, DocumentosIdentificacao, Endereco)


def normalizar_cargo(texto) -> str:
    """Chave de casamento do de-para de cargo: minúsculo, sem acento, espaços
    colapsados. "Analista DF Jr"/"analista df jr "/"Analista  DF  Jr" → mesma
    chave, mesmo tirvu_id."""
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFKD", str(texto or ""))
        if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", sem_acento).strip().lower()


# `tirvu_id_do_cargo` foi REMOVIDA em 2026-08-08: o Tirvu passou a casar o cargo
# pelo NOME, então o export escreve `c.cargo_funcao` direto. O de-para
# `CargoTirvu` continua existindo e sendo alimentado — ele guarda o CBO, que é o
# que distingue cargo homônimo ("AUXILIAR DE SERVIÇOS GERAIS" tem CBO de limpeza
# e de produção) —, só não é mais consultado aqui.

# A empregadora vai por RAZÃO SOCIAL desde 2026-08-08 (mesma virada de
# posto/cargo/jornada: o Tirvu passou a casar por texto). Antes ia o ID "1".
#
# O grupo opera com uma empregadora só (decisão do Bruno 2026-07-24), então o
# valor continua sendo um PADRÃO — não vira pendência nem depende de cadastro.
# Mas quem TEM `empresa_id` na ficha usa a razão social dela: se um dia houver
# uma segunda empregadora (Nossa Cozinha já foi citada), o export acerta sozinho
# em vez de carimbar Green House em todo mundo, que é o erro que não dá erro.
EMPRESA_RAZAO_SOCIAL_PADRAO = "GREEN HOUSE SERVICOS DE LOCACAO DE MAO DE OBRA LTDA"

COLUNAS_TIRVU = [
    "Empresa", "Posto de Serviço", "Matrícula", "Nome Completo", "CPF",
    "Cargo", "Data de Nascimento", "Data de Admissão", "Sexo (M ou F)",
    "Registra Ponto (S ou N)", "PIS", "CTPS Número", "CTPS Série", "Salário",
    "Salário - Complementar", "Salário - Extra", "Data Vigência - Salário",
    "Descrição da Jornada de Trabalho", "Whatsapp", "Últ. Período Aquisitivo",
    "Endereço", "Endereço - Número", "Endereço - Complemento",
    "Endereço - CEP", "Endereço - Bairro", "Endereço - Cidade",
    "Endereço - UF", "Login Sign-On",
]


def _so_digitos(v) -> str:
    return "".join(ch for ch in str(v or "") if ch.isdigit())


def cpf_mascarado(cpf) -> str:
    d = _so_digitos(cpf)
    if len(d) != 11:
        return str(cpf or "")
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"


def cep_mascarado(cep) -> str:
    """CEP no padrão do Tirvu: 00000-000 (ele exporta com hífen)."""
    d = _so_digitos(cep)
    if len(d) != 8:
        return str(cep or "")
    return f"{d[:5]}-{d[5:]}"


def ctps_do_cpf(cpf) -> tuple[str, str]:
    """CTPS Digital derivada do CPF, no formato que o Tirvu importa (feedback de
    campo 2026-07-24): NÚMERO = os 7 primeiros dígitos do CPF, SÉRIE = os 4
    últimos. Juntos reconstroem o CPF. NÃO é (CPF completo, "0000") — esse era o
    formato antigo, que o Tirvu recusou ("veio tudo junto, não separou os quatro
    últimos como série")."""
    d = _so_digitos(cpf)
    if len(d) != 11:
        return "", ""
    return d[:7], d[7:]


def _data(v) -> str:
    if v is None or v == "":
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v)  # Candidato guarda datas já como "dd/mm/aaaa"


def _salario(texto) -> object:
    """'R$ 1.500,00' / '1500' -> número (o Tirvu espera valor); se o texto não
    parsear, vai cru — melhor o Tirvu apontar do que sumir o dado."""
    if not texto:
        return ""
    limpo = re.sub(r"[Rr]\$|\s", "", str(texto))
    if re.fullmatch(r"\d{1,3}(\.\d{3})*(,\d{1,2})?|\d+(,\d{1,2})?", limpo):
        try:
            return float(limpo.replace(".", "").replace(",", "."))
        except ValueError:
            pass
    if re.fullmatch(r"\d+(\.\d{1,2})?", limpo):
        return float(limpo)
    return str(texto)


# Faixa ANTIGA (até 18/08/2026): `999` + 4 dígitos = 7 caracteres. Continua
# reconhecida — quem já a tem NÃO é renumerado: aquele número já foi para o
# Tirvu e para a planilha de ponto, e trocá-lo criaria duas matrículas para a
# mesma pessoa nos dois sistemas.
PREFIXO_MATRICULA_LEGADO = "999"
# Faixa NOVA (decisão do Bruno, 18/08/2026): `99` + 4 dígitos = **6 caracteres**,
# para caber no padrão de nome de arquivo `MATRÍCULA - NOME - DOCUMENTO` com
# matrícula de 6 posições. O prefixo se mantém (em vez de sequencial puro)
# porque é ele que distingue "matrícula que NÓS geramos" da que veio do Tirvu —
# `colaboradores.py` usa isso como indício antes de reverter um colaborador.
PREFIXO_MATRICULA_AUTO = "99"
_TAM_SEQ = 4


def _sequencial(matricula: str | None, prefixo: str, tamanho: int) -> int | None:
    """O sequencial de uma matrícula da faixa, ou `None` se ela não é da faixa.

    ⚠️ **A faixa `99` é ambígua com matrícula REAL do Tirvu.** `999001` tem 6
    dígitos e começa com `99`, então seria lida como sequencial 9001 — e o
    gerador pularia para `999002`, invadindo a numeração do Tirvu, onde outra
    pessoa pode já estar. Por isso a faixa nova exige que o dígito seguinte ao
    prefixo NÃO seja `9`: `990001`…`998999` são nossos, `999xxx` não é.
    Isso limita a faixa a ~9.000 matrículas, o que é folgado para uma base de
    ~1.200 pessoas e é o preço de não colidir com número de gente real.
    """
    d = _so_digitos(matricula)
    if len(d) != len(prefixo) + tamanho or not d.startswith(prefixo):
        return None
    resto = d[len(prefixo):]
    if prefixo == PREFIXO_MATRICULA_AUTO and resto[0] == "9":
        return None       # `999xxx` é matrícula do Tirvu, não da nossa faixa
    return int(resto)


def matricula_formatada(matricula: str | None, digitos: int = 6) -> str:
    """A matrícula com zeros à esquerda até `digitos`, para EXIBIR e NOMEAR.

    Pedido do Bruno (18/08/2026): nome de arquivo no padrão
    `MATRÍCULA - NOME - DOCUMENTO`, com a matrícula em 6 posições (`000000`).

    Duas coisas deliberadas:

    * **Não trunca o que é maior.** As matrículas da faixa legada têm 7 dígitos
      (`9990001`) e algumas do Tirvu podem ter mais — cortar para caber viraria
      OUTRO número, e o nome do arquivo deixaria de identificar a pessoa.
    * **Zero-pad é seguro para o casamento do ponto**: `import_ponto.matricula_norm`
      já ignora zeros à esquerda, então `003035` e `3035` continuam sendo a
      mesma pessoa. Isto formata para LER, não muda o que está gravado.
    """
    d = _so_digitos(matricula)
    if not d:
        return ""
    return d.zfill(digitos)


def matricula_automatica(matricula: str | None) -> bool:
    """Esta matrícula foi gerada por NÓS (qualquer uma das duas faixas)?

    Serve ao aviso de "esta pessoa provavelmente já está no Tirvu": só se gera
    matrícula automática no export, então tê-la é indício de que a linha já
    subiu para lá.
    """
    return any(_sequencial(matricula, p, _TAM_SEQ) is not None
               for p in (PREFIXO_MATRICULA_AUTO, PREFIXO_MATRICULA_LEGADO))


def proxima_matricula_auto(db: Session) -> str:
    """Próxima matrícula automática no padrão `99` + 4 dígitos (990001, 990002…).

    ⚠️ Considera as DUAS faixas ao decidir o próximo número. A antiga (`999`+4)
    não é mais gerada, mas segue existindo na base — e `9990007` e `990007`
    compartilham o mesmo sequencial. Ignorar a faixa antiga faria a primeira
    matrícula nova repetir um sequencial já usado, e as duas conviveriam na
    mesma coluna do export do Tirvu parecendo pessoas diferentes por um dígito.
    """
    from sqlalchemy import select
    matriculas = db.scalars(select(Candidato.matricula).where(
        Candidato.matricula.is_not(None))).all()
    maior = 0
    for m in matriculas:
        for prefixo in (PREFIXO_MATRICULA_AUTO, PREFIXO_MATRICULA_LEGADO):
            seq = _sequencial(m, prefixo, _TAM_SEQ)
            if seq is not None:
                maior = max(maior, seq)
                break
    return f"{PREFIXO_MATRICULA_AUTO}{maior + 1:0{_TAM_SEQ}d}"


def garantir_matricula(db: Session, c: Candidato) -> str:
    """Garante que o colaborador tenha matrícula: se não tiver, gera a automática
    (999+seq) e GRAVA no cadastro (fica estável para sempre). O caller faz o
    commit. O Tirvu exige matrícula — assim ninguém sobe sem ela."""
    if c.matricula and c.matricula.strip():
        return c.matricula
    nova = proxima_matricula_auto(db)
    c.matricula = nova
    db.flush()  # materializa para a próxima chamada na mesma transação não colidir
    return nova


def linha_tirvu(db: Session, c: Candidato, gerar_matricula: bool = False) -> dict:
    """Uma linha do layout, na ordem exata das 28 colunas.

    `gerar_matricula=True` (só no EXPORT) gera e GRAVA a matrícula automática
    quando faltar. Na pré-checagem de pendências fica False — consulta não muta
    dados (e a matrícula deixa de ser pendência, já que o export a gera)."""
    p = db.get(DadosPessoais, c.id)
    e = db.get(Endereco, c.id)
    d = db.get(DocumentosIdentificacao, c.id)
    posto = db.get(PostoServico, c.posto_servico_id) if c.posto_servico_id else None
    jornada = db.get(Jornada, c.jornada_id) if c.jornada_id else None
    # Empresa: usa a razão social da empregadora vinculada; sem vínculo (a
    # esmagadora maioria, já que o grupo tem uma só), cai no padrão.
    empresa = db.get(Empresa, c.empresa_id) if getattr(c, "empresa_id", None) else None
    razao_social = ((empresa.razao_social if empresa else "") or "").strip() \
        or EMPRESA_RAZAO_SOCIAL_PADRAO

    cpf = (d.cpf if d and d.cpf else c.cpf) or ""
    # CTPS para o Tirvu: SEMPRE derivar do CPF (número = 7 primeiros, série = 4
    # últimos). O que está GRAVADO em ctps_numero/serie pode ser o formato antigo
    # (CPF completo + "0000") de fichas anteriores a 2026-07-24 — o export não
    # backfilla o banco nem toca o PDF assinado, apenas gera a planilha certa.
    # Só cai no gravado se, por algum motivo, não houver CPF.
    ctps_num, ctps_serie = ctps_do_cpf(cpf) if cpf else ("", "")
    # Se o CPF é inválido/ausente (derivação vazia) mas há CTPS gravada, usa a
    # gravada — não perder o dado por causa de um CPF sujo (B1 da revisão).
    if not ctps_num and d and d.ctps_numero:
        ctps_num, ctps_serie = d.ctps_numero, d.ctps_serie or ""

    sexo = ""
    if p and p.sexo:
        sexo = "M" if p.sexo.value == "masculino" else "F"

    ponto = ""
    if c.registra_ponto is not None:
        ponto = "S" if c.registra_ponto else "N"

    nascimento = (p.data_nascimento if p and p.data_nascimento
                  else c.data_nascimento)

    # Endereço: coleta nova tem os campos separados; a antiga fica na string
    # única, que vai inteira na coluna "Endereço" (o Tirvu aceita as demais
    # vazias — validado pelo Bruno em 2026-07-19).
    logradouro, numero, complemento = "", "", ""
    if e:
        if e.logradouro:
            logradouro, numero, complemento = (
                e.logradouro, e.numero or "", e.complemento or "")
        else:
            logradouro = e.logradouro_numero_complemento or ""

    matricula = garantir_matricula(db, c) if gerar_matricula else (c.matricula or "")

    # Posto/Cargo/Jornada vão como TEXTO — e isto REVERTE a decisão de
    # 2026-07-24, de propósito.
    #
    # Naquela data o Tirvu casava por ID NUMÉRICO: colar o texto fazia ele
    # gravar zero, e por isso o export passou a escrever o `tirvu_id`. O Tirvu
    # MUDOU. O Bruno testou uma importação com o texto direto na célula em
    # 2026-08-08 e ela foi aceita — nas três colunas, inclusive Cargo, que agora
    # casa pelo NOME e não mais pelo id.
    #
    # A troca não é cosmética: com ID, quem estivesse num posto ou jornada sem
    # `tirvu_id` cadastrado saía com a célula VAZIA (19 postos e 23 jornadas da
    # base real estão nessa situação) e virava pendência. O texto é dado do
    # próprio cadastro e existe sempre — some a dependência do de-para.
    #
    # ⚠️ Se um dia o Tirvu voltar a exigir ID, o sintoma é traiçoeiro: ele
    # ACEITA a planilha e grava o vínculo ZERADO, sem reclamar. Não confie na
    # importação "ter passado" — confira um vínculo na tela do Tirvu.
    return {
        # Empresa vai por RAZÃO SOCIAL (nunca vazia: cai no padrão), então
        # continua fora de `pendencias_linha`.
        "Empresa": razao_social,
        "Posto de Serviço": (posto.nome if posto else "") or "",
        "Matrícula": matricula,
        "Nome Completo": c.nome_completo or "",
        "CPF": cpf_mascarado(cpf),
        # O Tirvu casa o cargo pelo NOME desde 2026-08-08 — o de-para
        # `CargoTirvu` deixou de ser necessário AQUI (segue vivo para o CBO,
        # que distingue homônimo na padronização).
        "Cargo": (c.cargo_funcao or "").strip(),
        "Data de Nascimento": _data(nascimento),
        "Data de Admissão": _data(c.data_admissao),
        "Sexo (M ou F)": sexo,
        "Registra Ponto (S ou N)": ponto,
        "PIS": (d.pis_nis_pasep if d else "") or "",
        "CTPS Número": ctps_num,
        "CTPS Série": ctps_serie,
        "Salário": _salario(c.salario_base),
        "Salário - Complementar": "",
        "Salário - Extra": "",
        "Data Vigência - Salário": "",
        # A coluna se chama "Descrição" e agora recebe a DESCRIÇÃO mesmo — que é
        # a forma canônica da jornada no cadastro (o texto único que o RH
        # confirma; os campos estruturados são metadados internos).
        "Descrição da Jornada de Trabalho": (jornada.descricao if jornada else "") or "",
        # só dígitos: o front agora guarda mascarado (61) 99999-8888, mas o
        # Tirvu recebe o telefone limpo (como PIS/CPF-sem-máscara alhures)
        "Whatsapp": _so_digitos(c.celular_whatsapp),
        "Últ. Período Aquisitivo": "",
        "Endereço": logradouro,
        "Endereço - Número": numero,
        "Endereço - Complemento": complemento,
        "Endereço - CEP": cep_mascarado(e.cep if e else ""),
        "Endereço - Bairro": (e.bairro if e else "") or "",
        "Endereço - Cidade": (e.cidade if e else "") or "",
        "Endereço - UF": (e.uf if e else "") or "",
        "Login Sign-On": "",
    }


ABA_TIRVU = "Plan1"


def montar_workbook_tirvu(linhas: list[dict]) -> bytes:
    """Gera a planilha EXATAMENTE no formato que o Tirvu aceita na importação:
    aba 'Plan1', as 28 colunas de COLUNAS_TIRVU em ordem FIXA (nunca a união das
    chaves), SEM auto-filtro, SEM painel congelado e SEM cabeçalho estilizado —
    o importador do Tirvu recusa planilhas com essa "decoração" (autoFilter no
    XML, aba com outro nome). Célula vazia é string vazia (não célula ausente/
    inlineStr solta), para o parser não tropeçar.

    Difere de propósito do `export_planilha.montar_workbook` (que é para o RH ler,
    com cor/filtro/congelamento) — este é para MÁQUINA, fiel ao modelo oficial
    `docs/Layout de Importação de Admissões.xlsx`."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = ABA_TIRVU
    # cabeçalho: texto puro, sem estilo (o modelo tem negrito, mas o que o Tirvu
    # lê é o TEXTO — mantemos simples e fiel à ordem)
    for j, nome in enumerate(COLUNAS_TIRVU, start=1):
        ws.cell(row=1, column=j, value=nome)
    for i, linha in enumerate(linhas, start=2):
        for j, nome in enumerate(COLUNAS_TIRVU, start=1):
            v = linha.get(nome, "")
            # Célula vazia: NÃO escreve (deixa ausente) — evita o
            # `<c t="inlineStr"></c>` malformado do openpyxl (tipo string sem o
            # elemento <is>), que faz parsers rígidos como o do Tirvu recusarem.
            # Só grava quando há conteúdo; números (salário) permanecem número.
            if v is None or v == "":
                continue
            ws.cell(row=i, column=j, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Rótulo amigável do campo na lista de pendências mostrada ao RH (a coluna do
# layout tem nome técnico; o RH lê o nome do campo como aparece na ficha).
_ROTULO_PENDENCIA = {
    # Desde 2026-08-08 estes saem como TEXTO (nome do posto, nome do cargo,
    # descrição da jornada): quando falta, o que falta é o VÍNCULO na ficha da
    # pessoa — não o cadastro de um ID. Dizer "ID Tirvu do posto" mandaria o RH
    # procurar no lugar errado.
    "Posto de Serviço": "Posto",
    "Cargo": "Cargo",
    "Descrição da Jornada de Trabalho": "Jornada",
    "Registra Ponto (S ou N)": "Registra Ponto",
}


def pendencias_linha(linha: dict) -> list[str]:
    """O que o Tirvu recusa/acusa como divergência no upload. Vai no aviso ao RH
    ANTES do upload — melhor saber aqui que descobrir na tela de divergências do
    Tirvu. A Matrícula NÃO entra: é auto-gerada (999+seq) no export. A Jornada
    entra: é dado real do cadastro que o Tirvu exige.

    "Registra Ponto" também entra (v1.82): é dado real do cadastro e sai em
    branco quando ninguém preencheu — o Tirvu aceita a célula vazia calado, e o
    colaborador nasce lá sem a marcação. Vira pendência aqui em vez de campo
    obrigatório no formulário: exigir na tela travaria a edição dos importados
    do Tirvu, que nasceram sem o campo."""
    faltas = []
    # "Empresa" não entra: é fixa (ID 1, Green House) — nunca falta.
    for campo in ("Nome Completo", "CPF", "PIS", "CTPS Número",
                  "Data de Admissão", "Posto de Serviço", "Cargo",
                  "Descrição da Jornada de Trabalho", "Registra Ponto (S ou N)"):
        if not linha.get(campo):
            faltas.append(_ROTULO_PENDENCIA.get(campo, campo))
    return faltas
