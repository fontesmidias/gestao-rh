"""Reembolso-Creche (IN SEGES/MGI 147/2026): página de acompanhamento do RH.

Nesta 1ª onda entrega o LEVANTAMENTO de elegibilidade por posto — a resposta
que os ofícios (CNMP nº 5/2026, ANATEL nº 45/2026) cobram em 5 dias úteis:
quantos colaboradores estão alocados em postos abrangidos pela IN. A camada de
dados das crianças (idade em anos/meses, documentos) entra na 2ª onda, quando o
autocadastro público estiver no ar; a estrutura já a comporta."""

import io
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.auth_rh import requer_rh
from app.core.db import get_db
from app.models.beneficio import BeneficioCreche, CriancaCreche, StatusBeneficio
from app.models.candidato import Candidato, PostoServico
from app.models.usuario_rh import UsuarioRH
from app.services import storage
from app.services.auditoria import registrar
from app.services.email_templates import enviar_modelo

router = APIRouter(tags=["creche-rh"], dependencies=[Depends(requer_rh)])


def _gerar_e_guardar_dossie(db: Session, ben: BeneficioCreche) -> str:
    from app.services.creche_pdf import gerar_dossie_creche
    pdf = gerar_dossie_creche(db, ben)
    key = f"creche/{ben.id}/dossie-reembolso-creche.pdf"
    storage.salvar(key, pdf, "application/pdf")
    ben.dossie_pdf_key = key
    ben.dossie_gerado_em = datetime.now(timezone.utc)
    return key


def partes_da_data(nasc: str | None) -> tuple[int, int, int] | None:
    """(dia, mês, ano) aceitando os DOIS formatos que existem no banco.

    Incidente de campo 2026-07-30: toda criança aparecia como "❌ passou de
    5a11m", inclusive um bebê de 2 anos. A causa era só isto — a função lia
    apenas `dd/mm/aaaa`, mas o `InputData.jsx` do wizard devolve **ISO**
    (`aaaa-mm-dd`) por padrão, e é assim que a maioria dos registros foi
    gravada. O `split("/")` falhava, a idade virava `None`, e `None` era tratado
    como "não elegível" — negando o benefício a quem tem direito.

    Aceitar os dois formatos é o certo aqui, e não "consertar os dados": o campo
    é `String(10)` livre e há registros antigos das duas formas. Uma migração
    cega reescreveria data de gente real com base em palpite de formato
    (`03/04` é 3 de abril ou 4 de março?), e isso decide dinheiro no
    contracheque.
    """
    if not nasc:
        return None
    texto = str(nasc).strip()
    try:
        if "-" in texto:                       # ISO: aaaa-mm-dd
            a, m, d = (int(x) for x in texto.split("-"))
        else:                                  # BR: dd/mm/aaaa
            d, m, a = (int(x) for x in texto.split("/"))
    except (ValueError, AttributeError):
        return None
    # Sanidade: data impossível não pode virar idade inventada.
    if not (1 <= m <= 12 and 1 <= d <= 31 and 1900 <= a <= 2200):
        return None
    return d, m, a


def data_br(nasc: str | None) -> str:
    """Exibição sempre em dd/mm/aaaa, venha o dado em que formato vier."""
    partes = partes_da_data(nasc)
    if partes is None:
        return nasc or "—"
    d, m, a = partes
    return f"{d:02d}/{m:02d}/{a:04d}"


def _idade_anos_meses(nasc: str, ref: datetime | None = None) -> tuple[int, int] | None:
    """Idade em (anos, meses). A IN 147 usa o limite de 5 anos e 11 meses — por
    isso os meses importam."""
    ref = ref or datetime.now(timezone.utc)
    partes = partes_da_data(nasc)
    if partes is None:
        return None
    d, m, a = partes
    anos = ref.year - a
    meses = ref.month - m
    if ref.day < d:
        meses -= 1
    if meses < 0:
        anos -= 1
        meses += 12
    return (anos, meses) if anos >= 0 else None


# Acima disto, a "criança" tem idade de adulto e o dado quase certamente é de
# outra pessoa. 18 é o limite de folga: cobre o caso real (o nascimento do PAI
# no campo do filho) sem chutar em cima da faixa do benefício, que é 5a11m.
IDADE_IMPLAUSIVEL_ANOS = 18


def _idade_implausivel(nasc: str, ref: datetime | None = None) -> bool:
    """A data é legível, mas não pode ser de uma criança deste benefício.

    Nasceu do caso real de 2026-08-02: a tela mostrou "Raul Moreira Monteiro ·
    12/10/1998 · 27a 9m · ❌ passou de 5a11m" para um filho que, na certidão,
    nasceu em 19/04/2022. A conta estava certíssima — o que estava no campo era
    o nascimento do PRÓPRIO COLABORADOR. O `InputData` não tem
    `autoComplete="off"`, então nem foi preciso erro humano: basta o navegador
    oferecer a data que a pessoa acabou de digitar em outro campo.

    Por que isto não é só cosmético: com 27 anos, `elegivel_idade` é False e
    `idade_desconhecida` é False — as duas condições que fazem `revisar_idade`
    disparar. Ou seja, o sistema marcava o benefício como RISCO DE GLOSA e
    empurrava o RH a suspender quem tinha direito. A defesa da v2.27 cobriu
    "data ilegível" e não previu "data legível e absurda".

    É AVISO, nunca bloqueio — a mesma escolha do `_indicio_tirvu`. Filho com
    deficiência não tem limite de idade em várias normas de benefício, e uma
    trava dura indeferiria calado um caso legítimo. Quem decide é o RH; o
    sistema só se recusa a fingir que leu um dado plausível.
    """
    am = _idade_anos_meses(nasc, ref)
    return am is not None and am[0] >= IDADE_IMPLAUSIVEL_ANOS


def _elegivel_por_idade(nasc: str, ref: datetime | None = None) -> bool:
    """<= 5 anos e 11 meses (art. 2º, §1º da IN 147/2026).

    `ref` existe para o teste fixar a data e não depender de quando roda — sem
    isso, um teste de idade passa hoje e falha no mês que vem.

    Data ilegível devolve False, mas quem chama precisa distinguir os dois casos
    na TELA: "passou da idade" e "não consegui ler a data" são coisas
    diferentes, e tratá-las igual foi o que fez o sistema negar benefício a
    quem tinha direito (ver `idade_desconhecida` em `_dump_crianca_rh`).
    """
    am = _idade_anos_meses(nasc, ref)
    if am is None:
        return False
    anos, meses = am
    return anos < 5 or (anos == 5 and meses <= 11)


def _fim_do_direito(nasc: str) -> str | None:
    """Último dia em que a criança ainda tem direito — o dia anterior aos 6 anos.

    A IN 147 dá direito até 5 anos e 11 meses, o que na prática significa "até
    a véspera do sexto aniversário". Devolver a DATA, e não só um sim/não, é o
    que transforma o benefício de reativo em previsível: o DP consegue saber em
    julho quem sai da folha em setembro, em vez de descobrir no fechamento.

    Devolve ISO (aaaa-mm-dd) — quem exibe usa `data_br`. `None` quando a data
    de nascimento não é legível: sem base não se inventa previsão.
    """
    partes = partes_da_data(nasc)
    if partes is None:
        return None
    d, m, a = partes
    # Seis anos depois, menos um dia. O 29/02 vira 28/02 no ano não bissexto:
    # `date` recusaria o dia inexistente, e adiar para 01/03 daria um dia a
    # mais de benefício do que a norma prevê.
    from datetime import date, timedelta
    try:
        seis = date(a + 6, m, d)
    except ValueError:
        seis = date(a + 6, m, 28) + timedelta(days=1)
    return (seis - timedelta(days=1)).isoformat()


def _postos_elegiveis(db: Session) -> list[PostoServico]:
    return db.scalars(
        select(PostoServico)
        .where(PostoServico.da_direito_creche == True)  # noqa: E712
        .order_by(PostoServico.nome)
    ).all()


@router.get("/rh/creche/resumo")
def resumo(db: Session = Depends(get_db)) -> dict:
    """Panorama do benefício: total de postos elegíveis e de colaboradores
    ativos alocados neles, quebrado por posto (com o valor de cada contrato)."""
    postos = _postos_elegiveis(db)
    ids = [p.id for p in postos]
    # contagem de colaboradores ATIVOS por posto elegível (uma consulta só)
    por_posto: dict = {pid: 0 for pid in ids}
    ativos = []
    if ids:
        ativos = db.scalars(
            select(Candidato).where(
                Candidato.posto_servico_id.in_(ids),
                Candidato.situacao == "ativo",
            )
        ).all()
        for c in ativos:
            por_posto[c.posto_servico_id] = por_posto.get(c.posto_servico_id, 0) + 1

    linhas = [{
        "posto_id": p.id, "posto": p.nome, "sigla": p.sigla,
        "contrato_ref": p.contrato_ref,
        "valor_reembolso": p.valor_reembolso_creche,
        "colaboradores_ativos": por_posto.get(p.id, 0),
    } for p in postos]

    # QUADRO DA CONSULTA (v2.34): quantos elegíveis, quantos se manifestaram e
    # quantos faltam. Antes o RH via o total de elegíveis num lugar e a lista de
    # pendentes em outro, sem nunca ver a conta fechar — e a pergunta do órgão
    # ("vocês consultaram todos?") não tinha uma resposta de uma linha.
    #
    # "Respondeu" inclui quem declarou que NÃO tem direito: a manifestação é o
    # que se prova, não o pedido. Um levantamento aberto e nunca enviado NÃO
    # conta — a pessoa entrou e parou no meio.
    ativos_ids = [c.id for c in ativos]
    respondidos = declararam_sem_direito = 0
    if ativos_ids:
        for b in db.scalars(select(BeneficioCreche).where(
                BeneficioCreche.candidato_id.in_(ativos_ids))).all():
            if b.status == StatusBeneficio.levantamento and b.enviado_em is None:
                continue
            respondidos += 1
            if b.status == StatusBeneficio.sem_direito_declarado:
                declararam_sem_direito += 1
    elegiveis = sum(por_posto.values())

    return {
        "postos_elegiveis": len(postos),
        "colaboradores_em_postos_elegiveis": elegiveis,
        "responderam": respondidos,
        "declararam_sem_direito": declararam_sem_direito,
        "faltam_responder": max(0, elegiveis - respondidos),
        "por_posto": linhas,
    }


@router.get("/rh/creche/pendentes-resposta")
def pendentes_resposta(db: Session = Depends(get_db),
                       _rh: UsuarioRH = Depends(requer_rh)) -> list[dict]:
    """Colaboradores ATIVOS em postos elegíveis que ainda NÃO responderam ao
    levantamento (não têm benefício, ou têm um em `levantamento` que nunca foi
    enviado). É a prova, para os órgãos (CNMP/ANATEL, prazo de 5 dias), de que
    todos os elegíveis foram consultados — e a lista de quem o RH precisa cobrar
    (auditoria 2026-07-22)."""
    postos = _postos_elegiveis(db)
    ids = [p.id for p in postos]
    if not ids:
        return []
    nomes_posto = {p.id: p.nome for p in postos}
    ativos = db.scalars(select(Candidato).where(
        Candidato.posto_servico_id.in_(ids),
        Candidato.situacao == "ativo")).all()
    # benefícios existentes por colaborador (o último estado conta como resposta,
    # exceto um levantamento que nunca foi enviado = não respondeu)
    bens = {b.candidato_id: b for b in db.scalars(select(BeneficioCreche)).all()}
    pendentes = []
    for c in ativos:
        b = bens.get(c.id)
        respondeu = b is not None and (
            b.status != StatusBeneficio.levantamento or b.enviado_em is not None)
        if not respondeu:
            pendentes.append({
                "candidato_id": c.id, "nome": c.nome_completo, "cpf": c.cpf,
                "matricula": c.matricula, "email": c.email,
                "posto": nomes_posto.get(c.posto_servico_id),
                "iniciou": b is not None,  # abriu o link mas não terminou
            })
    pendentes.sort(key=lambda x: (x["posto"] or "", x["nome"]))
    return pendentes


@router.get("/rh/creche/exportar")
def exportar(db: Session = Depends(get_db),
             rh: UsuarioRH = Depends(requer_rh)) -> Response:
    """Excel do levantamento: um colaborador ativo por linha, em postos que dão
    direito ao benefício, com o valor do reembolso do contrato. É a relação
    nominal que os órgãos pedem para instruir a repactuação."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    postos = {p.id: p for p in _postos_elegiveis(db)}
    colaboradores = []
    if postos:
        colaboradores = db.scalars(
            select(Candidato).where(
                Candidato.posto_servico_id.in_(list(postos.keys())),
                Candidato.situacao == "ativo",
            ).order_by(Candidato.nome_completo)
        ).all()

    cols = ["Nome completo", "CPF", "Matrícula", "Posto (contrato)", "Sigla",
            "Nº do contrato", "Valor do reembolso", "Data de admissão"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Elegiveis Reembolso-Creche"
    verde = PatternFill("solid", fgColor="0FB257")
    for j, nome in enumerate(cols, start=1):
        cel = ws.cell(row=1, column=j, value=nome)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = verde
        cel.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = max(14, min(40, len(nome) + 8))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for i, c in enumerate(colaboradores, start=2):
        p = postos.get(c.posto_servico_id)
        valores = [c.nome_completo, c.cpf, c.matricula,
                   p.nome if p else "", p.sigla if p else "",
                   p.contrato_ref if p else "",
                   p.valor_reembolso_creche if p else "", c.data_admissao]
        for j, v in enumerate(valores, start=1):
            ws.cell(row=i, column=j, value=v or "")

    buf = io.BytesIO()
    wb.save(buf)
    registrar(db, "creche_levantamento_exportado", ator="rh", ator_detalhe=rh.email,
              detalhe={"colaboradores": len(colaboradores), "postos": len(postos)})
    db.commit()
    agora = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="reembolso-creche-elegiveis-{agora}.xlsx"'},
    )


# ======================================================================
# Levantamentos: o RH revisa as adesões, aprova/ativa (com prazo mensal) ou
# indefere. Ao ativar, o colaborador recebe as orientações da entrega mensal.
# ======================================================================


def _dump_crianca_rh(c: CriancaCreche) -> dict:
    am = _idade_anos_meses(c.data_nascimento)
    return {
        "id": c.id, "nome": c.nome,
        # Sempre dd/mm/aaaa na tela, mesmo que o banco guarde ISO.
        "data_nascimento": data_br(c.data_nascimento),
        "parentesco": c.parentesco, "tipo_comprovante": c.tipo_comprovante,
        "idade_anos": am[0] if am else None, "idade_meses": am[1] if am else None,
        "elegivel_idade": _elegivel_por_idade(c.data_nascimento),
        # Data que não dá para ler NÃO é "fora da idade" — é dado a conferir.
        # Sem esta distinção, os dois casos apareciam como ❌ na tela e o RH
        # indeferiria por engano quem tem direito (incidente de 2026-07-30).
        "idade_desconhecida": am is None,
        # QUARTO estado: data legível, idade de adulto — quase sempre o
        # nascimento do colaborador digitado no campo do filho. Sem isto, o
        # caso aparece como um ❌ comum e o RH decide sobre dado errado.
        "idade_implausivel": _idade_implausivel(c.data_nascimento),
        "tem_certidao": bool(c.certidao_key), "tem_guarda": bool(c.guarda_key),
        # Decisão POR CRIANÇA (v2.55). `None` = ainda não decidida OU benefício
        # aprovado antes desta versão, quando a decisão era do conjunto — os
        # dois casos são legítimos e a tela os distingue pelo status.
        "decisao": c.decisao,
        "motivo_decisao": c.motivo_decisao,
        "decidido_por": c.decidido_por,
        "decidido_em": c.decidido_em,
    }


def _valor_unitario(v: str | None) -> float | None:
    """`R$ 526,64` → 526.64. `None` quando não dá para interpretar.

    Nunca devolve 0: valor ilegível tratado como zero entraria calado na conta
    do reembolso, e o total sairia menor sem nada acusar.
    """
    if not v:
        return None
    texto = str(v).strip().replace("R$", "").replace(" ", "")
    if not texto:
        return None
    try:
        return float(texto.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def _reais(v: float) -> str:
    """526.64 → `R$ 526,64` (pt-BR, com separador de milhar)."""
    inteiro, _, dec = f"{v:,.2f}".partition(".")
    return f"R$ {inteiro.replace(',', '.')},{dec}"


def _valor_total(ben: BeneficioCreche, criancas: list[dict]) -> str | None:
    """O que a pessoa recebe: valor unitário × crianças DEFERIDAS.

    O valor é por criança (decisão do Bruno, 2026-08-02), então indeferir uma
    reduz o total sozinho — sem o RH ter que recalcular à mão, que é onde o
    erro de folha aconteceria.

    Dois casos que NÃO se multiplicam, e os dois importam:

    * **Benefício anterior à v2.55** (todas as crianças com `decisao=None`): ali
      o valor gravado JÁ era o total do benefício. Multiplicá-lo agora dobraria
      o reembolso de quem tem dois filhos, calado, no contracheque.
    * **Valor ilegível**: devolve o texto cru, para o RH ver o que está gravado,
      em vez de um total inventado.
    """
    unit = _valor_unitario(ben.valor_reembolso)
    if unit is None:
        return ben.valor_reembolso
    decididas = [c for c in criancas if c["decisao"] is not None]
    if not decididas:
        return ben.valor_reembolso     # modelo antigo: o gravado é o total
    return _reais(unit * sum(1 for c in decididas if c["decisao"] == "deferida"))


def _dump_beneficio(db: Session, ben: BeneficioCreche) -> dict:
    col = db.get(Candidato, ben.candidato_id)
    posto = db.get(PostoServico, col.posto_servico_id) if col.posto_servico_id else None
    criancas = [_dump_crianca_rh(c) for c in ben.criancas]
    return {
        "id": ben.id, "candidato_id": col.id,
        "nome": col.nome_completo, "cpf": col.cpf, "matricula": col.matricula,
        "email": ben.email_confirmado or col.email, "telefone": ben.telefone,
        "posto": posto.nome if posto else None,
        "posto_da_direito": bool(posto and posto.da_direito_creche),
        "valor_posto": posto.valor_reembolso_creche if posto else None,
        "status": ben.status, "enviado_em": ben.enviado_em,
        "dia_entrega_mensal": ben.dia_entrega_mensal,
        "valor_reembolso": ben.valor_reembolso,
        "motivo_indeferimento": ben.motivo_indeferimento,
        "motivo_devolucao": ben.motivo_devolucao,
        "devolvido_em": ben.devolvido_em,
        # fila de acompanhamento (auditoria 2026-07-22): distingue "devolvi e
        # espero correção" de "colaborador só começou". aguardando_correcao =
        # voltou a levantamento por devolução e ainda não reenviou.
        "aguardando_correcao": (ben.status == StatusBeneficio.levantamento
                                and ben.devolvido_em is not None),
        "reenviado_apos_correcao": bool(ben.devolvido_em and ben.enviado_em
                                        and ben.enviado_em > ben.devolvido_em),
        "sem_direito_em": ben.sem_direito_em,
        "sem_direito_por": ben.sem_direito_por,
        "criancas": criancas,
        # ---- decisão por criança e o VALOR que ela determina (v2.55) -------
        # O `valor_reembolso` passou a ser UNITÁRIO, por criança deferida
        # (decisão do Bruno, 2026-08-02) — antes era o valor do benefício.
        # Quem já estava aprovado tem `decisao=None` em todas as crianças e é
        # tratado como "deferido pelo modelo anterior": aí o total é o próprio
        # valor gravado, sem multiplicar. Multiplicar retroativamente dobraria
        # o reembolso de quem tem dois filhos, em silêncio, no contracheque.
        "deferidas": sum(1 for c in criancas if c["decisao"] == "deferida"),
        "indeferidas": sum(1 for c in criancas if c["decisao"] == "indeferida"),
        "sem_decisao": sum(1 for c in criancas if c["decisao"] is None),
        "valor_unitario": ben.valor_reembolso,
        "valor_total": _valor_total(ben, criancas),
        "algum_elegivel": any(c["elegivel_idade"] for c in criancas),
        # alerta de idade (auditoria 2026-07-22): benefício ATIVO em que NENHUMA
        # criança ainda está na idade → o RH deve suspender (risco de glosa).
        # Só acusa quem REALMENTE passou da idade: criança com data ilegível
        # entra em `conferir_data`, não em risco de glosa — acusar as duas
        # coisas junto faria o RH suspender benefício de quem tem direito.
        # `idade_implausivel` fica de FORA do alarme pela mesma razão que
        # `idade_desconhecida`: nos dois casos não se sabe a idade real da
        # criança, e acusar risco de glosa faria o RH suspender benefício
        # legítimo — foi exatamente o que aconteceu no caso de 2026-08-02, em
        # que o dado do campo era o nascimento do próprio colaborador.
        "revisar_idade": (ben.status == StatusBeneficio.ativo and bool(criancas)
                          and not any(c["elegivel_idade"] for c in criancas)
                          and not any(c["idade_desconhecida"] for c in criancas)
                          and not any(c["idade_implausivel"] for c in criancas)),
        "conferir_data": any(c["idade_desconhecida"] or c["idade_implausivel"]
                             for c in criancas),
    }


@router.get("/rh/creche/levantamentos")
def listar_levantamentos(status: str | None = None,
                         db: Session = Depends(get_db)) -> list[dict]:
    """Adesões ao benefício. Por padrão as que precisam de ação (em análise);
    aceita filtro por status."""
    q = select(BeneficioCreche).order_by(BeneficioCreche.enviado_em.desc().nullslast())
    if status:
        q = q.where(BeneficioCreche.status == StatusBeneficio(status))
    return [_dump_beneficio(db, b) for b in db.scalars(q).all()]


# ROTA LITERAL antes da paramétrica `/{beneficio_id}`, senão "vigencia" seria
# lido como UUID e viraria 422 (armadilha registrada no CLAUDE.md).
@router.get("/rh/creche/vigencia")
def vigencia(db: Session = Depends(get_db)) -> dict:
    """Quem faz jus AGORA, quem deixou de fazer, e até quando cada um faz.

    Pedido do Bruno (2026-08-02): *"com base na data de nascimento da pessoa,
    quero que tenha um dash onde eu possa ver quem faz jus naquele momento,
    considerando a data atual e quem não faz mais e também ter a previsão de
    até quando a pessoa fará jus, e quando deixou de fazer. Isso vai ser
    importante pois o DP irá precisar saber mensalmente quem tem direito e não
    tem direito"*.

    A pergunta é mensal e recorrente, e a resposta é inteiramente DERIVADA da
    data de nascimento — nada aqui depende de coleta nova. O que o dash muda é
    o tempo do verbo: em vez de "esta criança está fora da idade" (constatação
    depois do fato), passa a responder "esta sai da folha em 12/09/2026"
    (previsão), que é o que permite ao DP se preparar em vez de corrigir.

    Uma linha por CRIANÇA, não por benefício: é a criança que faz aniversário,
    e um mesmo colaborador pode ter uma dentro e outra fora da idade — a linha
    por benefício esconderia exatamente o caso que exige decisão.

    Consultas em LOTE (benefícios → crianças → colaboradores → postos), nunca
    uma por linha: com a base inteira de elegíveis isso é a diferença entre
    abrir e travar.
    """
    hoje = datetime.now(timezone.utc)

    # Só benefícios que chegaram a ser decididos — quem está preenchendo ainda
    # não tem direito a apurar, e apareceria como "sem direito" por engano.
    bens = db.scalars(
        select(BeneficioCreche).where(
            BeneficioCreche.status.in_([StatusBeneficio.ativo,
                                        StatusBeneficio.aguardando_repactuacao]))
    ).all()
    if not bens:
        return {"gerado_em": hoje.isoformat(), "linhas": [],
                "resumo": {"com_direito": 0, "perderam": 0, "a_vencer_90d": 0,
                           "conferir": 0}}

    ids = [b.id for b in bens]
    criancas = db.scalars(
        select(CriancaCreche).where(CriancaCreche.beneficio_id.in_(ids))).all()
    col_ids = {b.candidato_id for b in bens}
    cols = {c.id: c for c in db.scalars(
        select(Candidato).where(Candidato.id.in_(col_ids))).all()}
    posto_ids = {c.posto_servico_id for c in cols.values() if c.posto_servico_id}
    postos = {p.id: p for p in db.scalars(
        select(PostoServico).where(PostoServico.id.in_(posto_ids))).all()} if posto_ids else {}
    por_beneficio = {b.id: b for b in bens}

    linhas = []
    for c in criancas:
        ben = por_beneficio.get(c.beneficio_id)
        if ben is None:
            continue
        col = cols.get(ben.candidato_id)
        if col is None:
            continue
        posto = postos.get(col.posto_servico_id) if col.posto_servico_id else None
        am = _idade_anos_meses(c.data_nascimento, hoje)
        fim = _fim_do_direito(c.data_nascimento)
        implausivel = _idade_implausivel(c.data_nascimento, hoje)
        tem_direito = _elegivel_por_idade(c.data_nascimento, hoje)

        # Quantos dias faltam para sair (negativo = já saiu). Serve para a
        # coluna "a vencer" e para ordenar por urgência.
        dias = None
        if fim is not None:
            from datetime import date
            dias = (date.fromisoformat(fim) - hoje.date()).days

        # A data de nascimento ilegível OU de adulto não vira "sem direito":
        # é dado a conferir. Tratar como negativa é o erro que já custou caro
        # duas vezes neste módulo.
        situacao = ("conferir" if (am is None or implausivel)
                    else "com_direito" if tem_direito else "perdeu")

        linhas.append({
            "crianca_id": c.id, "crianca": c.nome,
            "data_nascimento": data_br(c.data_nascimento),
            "idade": f"{am[0]}a {am[1]}m" if am else None,
            "parentesco": c.parentesco,
            "colaborador_id": col.id, "colaborador": col.nome_completo,
            "matricula": col.matricula, "cpf": col.cpf,
            "posto": posto.nome if posto else None,
            "beneficio_id": ben.id, "status_beneficio": ben.status,
            "valor_reembolso": ben.valor_reembolso or (
                posto.valor_reembolso_creche if posto else None),
            "situacao": situacao,
            # ISO nos dois: a tela formata, e o CSV precisa ser ordenável.
            "fim_do_direito": fim,
            "dias_para_o_fim": dias,
        })

    # Ordem: o que exige ação primeiro — quem está para sair, depois quem já
    # saiu, depois o resto.
    linhas.sort(key=lambda l: (l["dias_para_o_fim"] is None,
                               l["dias_para_o_fim"] if l["dias_para_o_fim"] is not None else 0))
    return {
        "gerado_em": hoje.isoformat(),
        "linhas": linhas,
        "resumo": {
            "com_direito": sum(1 for l in linhas if l["situacao"] == "com_direito"),
            "perderam": sum(1 for l in linhas if l["situacao"] == "perdeu"),
            "a_vencer_90d": sum(1 for l in linhas if l["situacao"] == "com_direito"
                                and l["dias_para_o_fim"] is not None
                                and 0 <= l["dias_para_o_fim"] <= 90),
            "conferir": sum(1 for l in linhas if l["situacao"] == "conferir"),
        },
    }


@router.get("/rh/creche/levantamentos/{beneficio_id}")
def detalhe_levantamento(beneficio_id: uuid.UUID, db: Session = Depends(get_db)) -> dict:
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    return _dump_beneficio(db, ben)


# rótulos amigáveis para o histórico (as ações de auditoria com prefixo creche_)
_HIST_ROTULO = {
    "creche_levantamento_enviado": "Colaborador enviou o levantamento",
    "creche_beneficio_devolvido": "RH devolveu para correção",
    "creche_beneficio_indeferido": "RH indeferiu",
    "creche_beneficio_ativado": "RH aprovou/ativou",
    "creche_beneficio_reaberto": "RH reabriu o levantamento",
    "creche_beneficio_suspenso": "RH suspendeu",
    "creche_beneficio_encerrado": "Benefício encerrado",
    "creche_sem_direito": "Registrado sem direito (não faz jus)",
    "creche_link_reenviado": "RH reenviou o link/código",
    "creche_codigo_enviado": "Código de acesso enviado",
    "creche_roteiro_falhou": "⚠️ Falha ao gerar o requerimento (reprocessar)",
}


@router.get("/rh/creche/levantamentos/{beneficio_id}/historico")
def historico_levantamento(beneficio_id: uuid.UUID, db: Session = Depends(get_db),
                           _rh: UsuarioRH = Depends(requer_rh)) -> list[dict]:
    """Linha do tempo das decisões do benefício (auditoria 2026-07-22): antes o
    RH só via o estado atual e o último revisor. Lê os eventos `creche_*` do
    colaborador, mais recente primeiro, com data/ator/motivo."""
    from app.models.evento import EventoAuditoria
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    eventos = db.scalars(
        select(EventoAuditoria)
        .where(EventoAuditoria.candidato_id == ben.candidato_id,
               EventoAuditoria.acao.like("creche\\_%", escape="\\"))
        .order_by(EventoAuditoria.criado_em.desc())).all()
    return [{
        "quando": e.criado_em, "ator": e.ator, "ator_detalhe": e.ator_detalhe,
        "acao": e.acao, "rotulo": _HIST_ROTULO.get(e.acao, e.acao),
        "motivo": (e.detalhe or {}).get("motivo") if e.detalhe else None,
    } for e in eventos]


@router.get("/rh/creche/tentativas-sem-acesso")
def tentativas_sem_acesso(db: Session = Depends(get_db),
                          _rh: UsuarioRH = Depends(requer_rh)) -> list[dict]:
    """Relatório das tentativas de acesso ao creche que NÃO resultaram em código
    enviado — para o RH distinguir "CPF realmente fora da base" de bug/dado
    errado (feedback 2026-07-27: colaboradores reais relataram 'CPF não está na
    base'). O gate público responde igual para todos (anti-enumeração), então
    ISTO é o único lugar onde o RH vê o que de fato aconteceu.

    TRÊS motivos:
    - `sem_match`: o CPF digitado NÃO casou com nenhum registro. Pode ser CPF
      realmente fora da base OU cadastrado errado/incompleto (ex.: zero à
      esquerda perdido na planilha, virou registro sem CPF). O RH confere.
    - `sem_email`: o CPF casou, mas o registro está SEM e-mail — a pessoa foi
      empurrada para a verificação por perguntas (KBA) e pode ter falhado. Basta
      o RH cadastrar o e-mail (aqui aparece nome e situação para localizar).
    - `codigo_recusado`: o código FOI enviado e mesmo assim a pessoa não
      entrou. Acrescentado em 2026-07-30 depois de uma colaboradora passar SEIS
      HORAS travada — sete códigos enviados com sucesso, nenhuma entrada — sem
      aparecer neste relatório, porque ele só enxergava quem nunca recebeu
      e-mail. Era o pior ponto cego: o envio funcionando dava a impressão de
      que estava tudo bem. Quem aparece aqui é candidato a "Reenviar link".

    Agrupa por CPF, com a contagem e a última tentativa (mais recente primeiro)."""
    from app.models.evento import EventoAuditoria
    eventos = db.scalars(
        select(EventoAuditoria)
        .where(EventoAuditoria.acao.in_(
            ["creche_iniciar_sem_match", "creche_iniciar_sem_email",
             "creche_codigo_recusado"]))
        .order_by(EventoAuditoria.criado_em.desc())).all()
    # Do mais fraco ao mais forte: um CPF que num momento não casou e depois
    # recebeu código e travou deve aparecer pelo problema MAIS ESPECÍFICO.
    peso = {"sem_match": 0, "sem_email": 1, "codigo_recusado": 2}
    motivo_de = {
        "creche_iniciar_sem_match": "sem_match",
        "creche_iniciar_sem_email": "sem_email",
        "creche_codigo_recusado": "codigo_recusado",
    }
    por_cpf: dict[str, dict] = {}
    for e in eventos:
        d = e.detalhe or {}
        cpf = d.get("cpf") or "—"
        motivo = motivo_de.get(e.acao, "sem_match")
        item = por_cpf.get(cpf)
        if item is None:
            por_cpf[cpf] = {
                "cpf": cpf, "motivo": motivo,
                "nome": d.get("nome"), "situacao": d.get("situacao"),
                "tentativas": 1, "ultima": e.criado_em, "primeira": e.criado_em,
            }
        else:
            item["tentativas"] += 1
            item["primeira"] = e.criado_em  # como vem desc, o último visto é o + antigo
            if peso[motivo] > peso[item["motivo"]]:
                item["motivo"] = motivo
            item["nome"] = item["nome"] or d.get("nome")
            item["situacao"] = item["situacao"] or d.get("situacao")
    return list(por_cpf.values())


class AtivarIn(BaseModel):
    dia_entrega_mensal: int | None = None
    valor_reembolso: str | None = None
    # se True, só aprova (aguardando_repactuacao); se False, ativa de fato
    aguardar_repactuacao: bool = False


@router.post("/rh/creche/levantamentos/{beneficio_id}/ativar")
def ativar_beneficio(beneficio_id: uuid.UUID, payload: AtivarIn, db: Session = Depends(get_db),
                     rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Aprova o benefício. Se aguardar_repactuacao=True, fica em
    'aguardando_repactuacao'; senão vai a 'ativo' e o colaborador recebe as
    orientações da entrega mensal (com o prazo)."""
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    col = db.get(Candidato, ben.candidato_id)
    posto = db.get(PostoServico, col.posto_servico_id) if col.posto_servico_id else None

    # ---- decisão por criança (v2.55) --------------------------------------
    # Se o RH decidiu ALGUMA criança, ele tem que ter decidido TODAS: aprovar
    # com uma pendente deixaria no requerimento um dependente sem análise, e o
    # valor sairia errado (o reembolso é por criança deferida).
    #
    # Quem não usou a decisão individual segue pelo caminho antigo — aprovar o
    # conjunto —, que é o que mantém compatível o benefício aberto antes desta
    # versão.
    criancas = list(ben.criancas)
    decididas = [c for c in criancas if c.decisao is not None]
    if decididas and len(decididas) != len(criancas):
        pendentes = [c.nome for c in criancas if c.decisao is None]
        raise HTTPException(status_code=409,
                            detail={"erro": "criancas_sem_decisao",
                                    "criancas": pendentes})
    # Todas negadas: não há o que reembolsar. Vira indeferimento, com os
    # motivos agregados — em vez de um benefício "ativo" que paga zero, que
    # seria mentira no relatório e no requerimento.
    if decididas and all(c.decisao == "indeferida" for c in criancas):
        ben.status = StatusBeneficio.indeferido
        ben.motivo_indeferimento = "; ".join(
            f"{c.nome}: {c.motivo_decisao or 'sem motivo registrado'}" for c in criancas)
        ben.revisado_por, ben.revisado_em = rh.email, datetime.now(timezone.utc)
        registrar(db, "creche_beneficio_indeferido", ator="rh", ator_detalhe=rh.email,
                  candidato_id=col.id,
                  detalhe={"motivo": ben.motivo_indeferimento,
                           "por_crianca": True})
        db.commit()
        try:
            # Lê o motivo do próprio benefício (já gravado acima), como o
            # indeferimento normal faz — o e-mail sai igual, seja a negativa do
            # conjunto ou a soma das negativas por criança.
            _email_indeferimento(db, ben, col)
        except Exception:
            pass
        return _dump_beneficio(db, ben)

    if payload.dia_entrega_mensal is not None:
        ben.dia_entrega_mensal = max(1, min(28, payload.dia_entrega_mensal))
    ben.valor_reembolso = (payload.valor_reembolso
                           or (posto.valor_reembolso_creche if posto else None))
    ben.revisado_por = rh.email
    ben.revisado_em = datetime.now(timezone.utc)
    # gera o dossiê do benefício (requerimento + anexos + declaração-modelo)
    try:
        _gerar_e_guardar_dossie(db, ben)
    except Exception:
        pass  # o dossiê é reproduzível pelo botão; não trava a ativação
    if payload.aguardar_repactuacao:
        ben.status = StatusBeneficio.aguardando_repactuacao
    else:
        ben.status = StatusBeneficio.ativo
        ben.ativado_em = datetime.now(timezone.utc)
        # roteiro de assinatura do requerimento: colaborador (na sessão de creche)
        # → RH que aprovou. Só após a aprovação é que o colaborador pode assinar.
        from app.services.roteiro_assinatura import criar_roteiro_creche
        try:
            criar_roteiro_creche(db, ben, rh)
        except Exception:
            # NÃO engolir sem rastro: sem o roteiro o colaborador nunca vê o botão
            # de assinar. Registra para o RH reprocessar (auditoria 2026-07-22).
            registrar(db, "creche_roteiro_falhou", ator="sistema",
                      candidato_id=col.id, detalhe={"beneficio": str(ben.id)})
    registrar(db, "creche_beneficio_ativado", ator="rh", ator_detalhe=rh.email,
              candidato_id=col.id,
              detalhe={"status": ben.status.value, "dia": ben.dia_entrega_mensal})
    db.commit()
    # e-mails após o commit (SMTP fora não desfaz a decisão)
    try:
        if ben.status == StatusBeneficio.ativo:
            _email_orientacoes_mensais(db, ben, col)
        else:
            _email_aguardando_repactuacao(db, ben, col)
    except Exception:
        pass
    return _dump_beneficio(db, ben)


class DecisaoCriancaIn(BaseModel):
    decisao: str                  # deferida | indeferida
    motivo: str | None = None


@router.post("/rh/creche/levantamentos/{beneficio_id}/criancas/{crianca_id}/decidir")
def decidir_crianca(beneficio_id: uuid.UUID, crianca_id: uuid.UUID,
                    payload: DecisaoCriancaIn, db: Session = Depends(get_db),
                    rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Defere ou indefere UMA criança, sem decidir o benefício inteiro.

    Feedback 2026-08-02: *"se a pessoa tem mais de um filho e um eu defiro e
    outro eu indefiro, não tem opção individual por filho, somente indeferir
    tudo ou aprovar tudo, não tá legal isso"*.

    Antes desta rota, o único caminho para negar uma criança era DEVOLVER o
    levantamento e pedir que o colaborador a removesse — o que apagava a prova
    de que ela havia sido analisada e negada. O registro do indeferimento é
    justamente o que demonstra que o RH avaliou aquele dependente.

    Continua sendo **um único requerimento**: a decisão por criança alimenta o
    mesmo PDF, que lista as deferidas e, em seção própria, as negadas com o
    motivo.
    """
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    crianca = db.get(CriancaCreche, crianca_id)
    # A criança TEM que pertencer a este benefício: sem esta checagem, o id de
    # uma criança de outra família seria aceito (mesma guarda do
    # `baixar_doc_crianca`).
    if crianca is None or crianca.beneficio_id != ben.id:
        raise HTTPException(status_code=404, detail="crianca_nao_encontrada")
    if payload.decisao not in ("deferida", "indeferida"):
        raise HTTPException(status_code=422, detail="decisao_invalida")
    # Motivo obrigatório para negar: é o que o colaborador vê e o que sustenta
    # a decisão numa eventual contestação. Deferir não precisa de justificativa.
    motivo = (payload.motivo or "").strip()
    if payload.decisao == "indeferida" and not motivo:
        raise HTTPException(status_code=422, detail="motivo_obrigatorio")
    # Depois de encerrado não se redecide: reabra o levantamento antes.
    if ben.status in (StatusBeneficio.encerrado, StatusBeneficio.suspenso):
        raise HTTPException(status_code=409, detail="beneficio_encerrado")

    crianca.decisao = payload.decisao
    crianca.motivo_decisao = motivo or None
    crianca.decidido_por = rh.email
    crianca.decidido_em = datetime.now(timezone.utc)
    registrar(db, "creche_crianca_decidida", ator="rh", ator_detalhe=rh.email,
              candidato_id=ben.candidato_id,
              detalhe={"crianca": crianca.nome, "decisao": payload.decisao,
                       "motivo": motivo or None})
    db.commit()
    return _dump_beneficio(db, ben)


class IndeferirIn(BaseModel):
    motivo: str


@router.post("/rh/creche/levantamentos/{beneficio_id}/indeferir")
def indeferir_beneficio(beneficio_id: uuid.UUID, payload: IndeferirIn,
                        db: Session = Depends(get_db),
                        rh: UsuarioRH = Depends(requer_rh)) -> dict:
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    ben.status = StatusBeneficio.indeferido
    ben.motivo_indeferimento = payload.motivo.strip() or None
    ben.revisado_por = rh.email
    ben.revisado_em = datetime.now(timezone.utc)
    col = db.get(Candidato, ben.candidato_id)
    registrar(db, "creche_beneficio_indeferido", ator="rh", ator_detalhe=rh.email,
              candidato_id=ben.candidato_id, detalhe={"motivo": ben.motivo_indeferimento})
    db.commit()
    try:
        _email_indeferimento(db, ben, col)  # avisa o colaborador (não trava)
    except Exception:
        pass
    return _dump_beneficio(db, ben)


class DevolverIn(BaseModel):
    motivo: str


@router.post("/rh/creche/levantamentos/{beneficio_id}/devolver")
def devolver_beneficio(beneficio_id: uuid.UUID, payload: DevolverIn,
                       db: Session = Depends(get_db),
                       rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Devolve o levantamento ao colaborador para correção (feedback
    2026-07-21). O status volta a `levantamento` — o que reabre a edição no link
    público e permite reenviar — com um motivo VISÍVEL ao colaborador. Limpa o
    envio anterior e um eventual indeferimento (a devolução é uma segunda
    chance, não um veredito)."""
    if not (payload.motivo or "").strip():
        raise HTTPException(status_code=422, detail="motivo_obrigatorio")
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    # guard: devolver só faz sentido para pedido em análise/aguardando. Sem isso
    # um clique fora de ordem devolveria um ATIVO, reabrindo edição de benefício
    # que já tem dossiê/assinatura (deixaria artefatos órfãos). Reabrir um
    # terminal (indeferido/sem-direito) é a rota /reabrir, não esta.
    if ben.status not in (StatusBeneficio.em_analise, StatusBeneficio.aguardando_repactuacao):
        raise HTTPException(status_code=409, detail="nao_devolvivel")
    ben.status = StatusBeneficio.levantamento
    ben.motivo_devolucao = payload.motivo.strip()
    ben.devolvido_em = datetime.now(timezone.utc)
    ben.motivo_indeferimento = None  # devolver anula um indeferimento anterior
    ben.enviado_em = None            # o colaborador vai reenviar
    ben.dados_conferidos_em = None   # e reconferir os dados
    ben.revisado_por = rh.email
    ben.revisado_em = datetime.now(timezone.utc)
    col = db.get(Candidato, ben.candidato_id)
    # Link de acesso DIRETO no e-mail (v1.82): o e-mail já é comprovado, então
    # refazer o 2FA só para corrigir um dado é atrito que faz a correção não
    # voltar. Emitido antes do commit para entrar na mesma transação.
    from app.api.creche_publico import emitir_acesso_devolucao
    token = emitir_acesso_devolucao(db, ben)
    registrar(db, "creche_beneficio_devolvido", ator="rh", ator_detalhe=rh.email,
              candidato_id=ben.candidato_id, detalhe={"motivo": ben.motivo_devolucao})
    db.commit()
    try:
        _email_devolucao(db, ben, col, token)  # avisa e já leva o link de correção
    except Exception:
        pass
    return _dump_beneficio(db, ben)


class ReenviarLinkIn(BaseModel):
    email: str | None = None  # se vier, corrige o e-mail do colaborador antes


@router.post("/rh/creche/levantamentos/{beneficio_id}/reenviar-link")
def reenviar_link_creche(beneficio_id: uuid.UUID, payload: ReenviarLinkIn,
                         db: Session = Depends(get_db),
                         rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Destrava o colaborador que não consegue entrar (feedback 2026-07-22): e-mail
    não chegou, código expirou, ou o e-mail na base está errado. O RH pode
    corrigir o e-mail (com auditoria) e reenvia o código 2FA. Sem e-mail, não há
    como enviar — devolve 422 para o RH resolver o contato antes."""
    from app.api.creche_publico import _gerar_e_enviar_codigo
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    col = db.get(Candidato, ben.candidato_id)
    novo_email = (payload.email or "").strip()
    if novo_email:
        antes = col.email
        col.email = novo_email
        ben.email_confirmado = None  # o novo e-mail passará pelo 2FA de novo
        registrar(db, "creche_email_corrigido", ator="rh", ator_detalhe=rh.email,
                  candidato_id=col.id, detalhe={"antes": antes, "depois": novo_email})
        db.commit()
    destino = ben.email_confirmado or col.email
    if not destino:
        raise HTTPException(status_code=422, detail="sem_email")
    _gerar_e_enviar_codigo(db, col, ben, destino)
    registrar(db, "creche_link_reenviado", ator="rh", ator_detalhe=rh.email,
              candidato_id=col.id)
    db.commit()
    return {"enviado_para": destino}


def _marcar_sem_direito(db: Session, ben: BeneficioCreche, por: str) -> None:
    """Marca a declaração de que o colaborador não tem dependentes que dão
    direito. `por` = 'colaborador' (declarou no link) ou o e-mail do RH."""
    ben.status = StatusBeneficio.sem_direito_declarado
    ben.sem_direito_em = datetime.now(timezone.utc)
    ben.sem_direito_por = por


@router.post("/rh/creche/colaboradores/{colaborador_id}/sem-direito")
def rh_marcar_sem_direito(colaborador_id: uuid.UUID, db: Session = Depends(get_db),
                          rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """O RH registra que o colaborador (elegível por posto) declarou não ter
    dependentes que dão direito — para quem respondeu por fora (WhatsApp,
    pessoalmente). Cria o benefício se ainda não existir. Fica no relatório
    como 'consultado e não pediu' (feedback 2026-07-21)."""
    col = db.get(Candidato, colaborador_id)
    if col is None:
        raise HTTPException(status_code=404, detail="colaborador_nao_encontrado")
    ben = db.scalar(select(BeneficioCreche)
                    .where(BeneficioCreche.candidato_id == colaborador_id))
    if ben is None:
        ben = BeneficioCreche(candidato_id=colaborador_id)
        db.add(ben)
    elif ben.status == StatusBeneficio.ativo:
        # não apaga um benefício em pagamento por engano de clique
        raise HTTPException(status_code=409, detail="beneficio_ativo")
    _marcar_sem_direito(db, ben, rh.email)
    registrar(db, "creche_sem_direito", ator="rh", ator_detalhe=rh.email,
              candidato_id=colaborador_id, detalhe={"por": "rh"})
    db.commit()
    try:
        _email_sem_direito(db, ben, col)  # confirmação escrita ao colaborador
    except Exception:
        pass
    return _dump_beneficio(db, ben)


@router.post("/rh/creche/levantamentos/{beneficio_id}/reabrir")
def reabrir_beneficio(beneficio_id: uuid.UUID, db: Session = Depends(get_db),
                      rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Devolve o benefício a `levantamento` para o colaborador refazer/completar.

    Cobre três casos (feedback 2026-07-22):
    - indeferido por engano (ou o colaborador quer corrigir e reapresentar);
    - quem declarou 'sem direito' passou a ter filho/guarda;
    - **MAIS FILHOS num benefício ATIVO**: o modelo é 1 benefício : N crianças,
      então "novo requerimento" se resolve reabrindo e ACRESCENTANDO a criança —
      sem duplicar benefício (decisão do Bruno 2026-07-22, evita a migração 1:N).
      Reabrir um ativo o tira do pagamento até o RH aprovar de novo, por isso
      fica explícito na auditoria.
    Encerrado/suspenso não se reabre por aqui (o vínculo/idade mudou: é caso de
    novo levantamento pelo link, ou de reativar via aprovação)."""
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    if ben.status not in (StatusBeneficio.indeferido,
                          StatusBeneficio.sem_direito_declarado,
                          StatusBeneficio.ativo,
                          StatusBeneficio.aguardando_repactuacao):
        raise HTTPException(status_code=409, detail="nao_reabrivel")
    era_ativo = ben.status in (StatusBeneficio.ativo, StatusBeneficio.aguardando_repactuacao)
    ben.status = StatusBeneficio.levantamento
    ben.motivo_indeferimento = None
    ben.sem_direito_em = None
    ben.sem_direito_por = None
    ben.enviado_em = None
    ben.dados_conferidos_em = None
    ben.revisado_por = rh.email
    ben.revisado_em = datetime.now(timezone.utc)
    col = db.get(Candidato, ben.candidato_id)
    registrar(db, "creche_beneficio_reaberto", ator="rh", ator_detalhe=rh.email,
              candidato_id=ben.candidato_id,
              detalhe={"era_ativo": era_ativo,
                       "motivo": "inclusão de criança" if era_ativo else "reabertura"})
    db.commit()
    if era_ativo:
        try:
            _email_reabertura_para_incluir(db, ben, col)
        except Exception:
            pass
    return _dump_beneficio(db, ben)


class EncerrarIn(BaseModel):
    motivo: str
    encerrar: bool = False  # False = suspender (reversível), True = encerrar


@router.post("/rh/creche/levantamentos/{beneficio_id}/suspender")
def suspender_beneficio(beneficio_id: uuid.UUID, payload: EncerrarIn,
                        db: Session = Depends(get_db),
                        rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Tira um benefício ATIVO de circulação: suspende (criança passou de 5a11m,
    pendência) ou encerra (desligamento). Para o ciclo mensal e avisa o
    colaborador — sem isso o RH seguia orientado a reembolsar quem já não tem
    direito (risco de glosa na prestação de contas, auditoria 2026-07-22)."""
    if not (payload.motivo or "").strip():
        raise HTTPException(status_code=422, detail="motivo_obrigatorio")
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    if ben.status not in (StatusBeneficio.ativo, StatusBeneficio.aguardando_repactuacao):
        raise HTTPException(status_code=409, detail="nao_suspensivel")
    ben.status = StatusBeneficio.encerrado if payload.encerrar else StatusBeneficio.suspenso
    ben.motivo_indeferimento = payload.motivo.strip()  # reusa o campo de motivo
    ben.revisado_por = rh.email
    ben.revisado_em = datetime.now(timezone.utc)
    col = db.get(Candidato, ben.candidato_id)
    registrar(db, "creche_beneficio_encerrado" if payload.encerrar else "creche_beneficio_suspenso",
              ator="rh", ator_detalhe=rh.email, candidato_id=col.id,
              detalhe={"motivo": payload.motivo.strip()})
    db.commit()
    try:
        _email_suspensao(db, ben, col, payload.motivo.strip(), payload.encerrar)
    except Exception:
        pass
    return _dump_beneficio(db, ben)


class PrazoMassaIn(BaseModel):
    beneficio_ids: list[uuid.UUID]
    dia_entrega_mensal: int


@router.put("/rh/creche/prazos")
def editar_prazos(payload: PrazoMassaIn, db: Session = Depends(get_db),
                  rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Ajusta o dia de entrega mensal de vários benefícios de uma vez (ou de um,
    passando um id só)."""
    dia = max(1, min(28, payload.dia_entrega_mensal))
    bens = db.scalars(select(BeneficioCreche)
                      .where(BeneficioCreche.id.in_(payload.beneficio_ids))).all()
    for b in bens:
        b.dia_entrega_mensal = dia
    registrar(db, "creche_prazos_alterados", ator="rh", ator_detalhe=rh.email,
              detalhe={"qtd": len(bens), "dia": dia})
    db.commit()
    return {"atualizados": len(bens), "dia_entrega_mensal": dia}


class CondicoesIn(BaseModel):
    """Prazo e valor de UM benefício, editáveis depois de aprovado."""
    dia_entrega_mensal: int | None = None
    valor_reembolso: str | None = None
    motivo: str | None = None


@router.put("/rh/creche/levantamentos/{beneficio_id}/condicoes")
def editar_condicoes(beneficio_id: uuid.UUID, payload: CondicoesIn,
                     db: Session = Depends(get_db),
                     rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Corrige o dia de entrega e o VALOR de um benefício já aprovado.

    Pedido do Bruno (2026-08-02): *"para os reembolso creche que já obtiveram a
    aprovação e estão aguardando a repactuação, quero que seja possível editar
    ali no painel, tanto a data limite para ser remetida a documentação mensal
    pelo colaborador, quanto também o valor do reembolso"*.

    Faltava mesmo: o `valor_reembolso` só era gravado dentro de
    `ativar_beneficio`, então repactuar um contrato deixava os benefícios já
    ativos com o valor antigo congelado — e o único jeito de mexer era
    RE-ATIVAR, o que regenera dossiê, recria roteiro de assinatura e dispara
    e-mail para o colaborador. Muito estrago para trocar um número.

    O valor NÃO é propagado do posto automaticamente, de propósito: o campo do
    benefício existe justamente porque ele pode divergir do contrato (é uma
    cópia congelada na ativação). Quem decide se a repactuação vale para um
    caso específico é o RH.

    Campo ausente no payload = não mexe. Assim dá para corrigir só o valor sem
    ter que reenviar o prazo, e vice-versa.
    """
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    # Só faz sentido em benefício que JÁ foi decidido — antes disso, a condição
    # é definida na própria aprovação, e editar aqui criaria dois caminhos para
    # a mesma coisa.
    if ben.status not in (StatusBeneficio.ativo, StatusBeneficio.aguardando_repactuacao):
        raise HTTPException(status_code=409, detail="beneficio_nao_aprovado")

    antes = {"dia_entrega_mensal": ben.dia_entrega_mensal,
             "valor_reembolso": ben.valor_reembolso}
    if payload.dia_entrega_mensal is not None:
        # 1-28 como no `editar_prazos`: dia 29-31 não existe em todo mês, e um
        # prazo que some em fevereiro é prazo que o colaborador perde.
        ben.dia_entrega_mensal = max(1, min(28, payload.dia_entrega_mensal))
    if payload.valor_reembolso is not None:
        novo = payload.valor_reembolso.strip()
        ben.valor_reembolso = novo or None

    registrar(db, "creche_condicoes_alteradas", ator="rh", ator_detalhe=rh.email,
              candidato_id=ben.candidato_id,
              detalhe={"antes": antes,
                       "depois": {"dia_entrega_mensal": ben.dia_entrega_mensal,
                                  "valor_reembolso": ben.valor_reembolso},
                       "motivo": (payload.motivo or "").strip() or None})
    db.commit()
    return _dump_beneficio(db, ben)


def _email_orientacoes_mensais(db: Session, ben: BeneficioCreche,
                               col: Candidato) -> None:
    """Enviado ao ATIVAR: orienta a entrega mensal da documentação de despesa."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    enviar_modelo(db, "creche_ativado", email, {
        "nome": col.nome_completo.split()[0].title(),
        "dia": ben.dia_entrega_mensal,
    })


def _url_creche() -> str:
    from app.core.config import get_settings
    return f"{get_settings().base_url.rstrip('/')}/creche"


def _email_devolucao(db: Session, ben: BeneficioCreche, col: Candidato,
                     token: str | None = None) -> None:
    """Avisa o colaborador que o RH devolveu o levantamento para correção — sem
    isso ele só descobriria se reabrisse o link por acaso (feedback 2026-07-22).

    Vai com LINK DIRETO (v1.82, pedido do Bruno): quem foi devolvido já validou
    o e-mail alguma vez, então refazer o 2FA só para corrigir um dado faz a
    correção não voltar. O token vale 7 dias e a devolução seguinte derruba o
    anterior. Sem token (chamada antiga), cai no texto de sempre — entrar pelo
    CPF continua funcionando."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    # O link vai no BOTÃO (v2.06): antes o texto tinha dois ramos inteiros só
    # para dizer "com token" ou "sem token" — agora a diferença é só a URL.
    url = _url_creche()
    enviar_modelo(db, "creche_devolvido", email, {
        "nome": col.nome_completo.split()[0].title(),
        "motivo": ben.motivo_devolucao or "verifique os dados e reenvie.",
        "link": f"{url}?t={token}" if token else url,
    })


def _email_aguardando_repactuacao(db: Session, ben: BeneficioCreche,
                                  col: Candidato) -> None:
    """Avisa o colaborador de que foi APROVADO, mas o pagamento depende do ajuste
    (repactuação) do contrato do posto — senão ele acha que ainda está 'em
    análise' e cobra o RH sem necessidade (auditoria 2026-07-22)."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    enviar_modelo(db, "creche_aguardando_contrato", email, {
        "nome": col.nome_completo.split()[0].title()})


def encerrar_creche_no_desligamento(db: Session, candidato_id) -> None:
    """Gancho chamado quando o colaborador é desligado: encerra o benefício
    creche ativo/aguardando (não faz sentido reembolsar quem saiu). Idempotente
    e silencioso — não trava o desligamento. NÃO faz commit (o chamador commita).
    Avisa o colaborador por e-mail."""
    ben = db.scalar(select(BeneficioCreche).where(
        BeneficioCreche.candidato_id == candidato_id,
        BeneficioCreche.status.in_((StatusBeneficio.ativo,
                                    StatusBeneficio.aguardando_repactuacao))))
    if ben is None:
        return
    ben.status = StatusBeneficio.encerrado
    ben.motivo_indeferimento = "Colaborador desligado"
    col = db.get(Candidato, candidato_id)
    registrar(db, "creche_beneficio_encerrado", ator="sistema",
              candidato_id=candidato_id, detalhe={"motivo": "desligamento"})
    try:
        _email_suspensao(db, ben, col, "Colaborador desligado", encerrado=True)
    except Exception:
        pass


def _email_reabertura_para_incluir(db: Session, ben: BeneficioCreche,
                                   col: Candidato) -> None:
    """Reabertura de um benefício ATIVO para o colaborador acrescentar outra
    criança (mais filhos). O benefício fica fora do pagamento até a nova
    aprovação — por isso o pedido de urgência (2026-07-22)."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    enviar_modelo(db, "creche_incluir_crianca", email, {
        "nome": col.nome_completo.split()[0].title(), "link": _url_creche()})


def _email_sem_direito(db: Session, ben: BeneficioCreche, col: Candidato) -> None:
    """Quando o RH registra 'sem direito' por um colaborador (respondeu por fora),
    manda uma confirmação escrita — ele pode contestar antes de virar relatório
    oficial, e a trilha de auditoria fica mais forte (auditoria 2026-07-22)."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    enviar_modelo(db, "creche_sem_direito", email, {
        "nome": col.nome_completo.split()[0].title()})


def _email_suspensao(db: Session, ben: BeneficioCreche, col: Candidato,
                     motivo: str, encerrado: bool) -> None:
    """Avisa o colaborador de que o benefício foi suspenso/encerrado e que ele
    NÃO precisa mais enviar a comprovação mensal (auditoria 2026-07-22)."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    enviar_modelo(db, "creche_suspenso", email, {
        "nome": col.nome_completo.split()[0].title(),
        "verbo": "encerrado" if encerrado else "suspenso",
        "motivo": motivo,
    })


def _email_indeferimento(db: Session, ben: BeneficioCreche, col: Candidato) -> None:
    """Avisa o colaborador do indeferimento com o motivo (antes: silencioso)."""
    email = ben.email_confirmado or col.email
    if not email:
        return
    enviar_modelo(db, "creche_indeferido", email, {
        "nome": col.nome_completo.split()[0].title(),
        "motivo": ben.motivo_indeferimento or "não atende aos requisitos do benefício.",
    })


@router.post("/rh/creche/levantamentos/{beneficio_id}/dossie")
def gerar_dossie_endpoint(beneficio_id: uuid.UUID, db: Session = Depends(get_db),
                          rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """(Re)gera o dossiê do benefício sob demanda."""
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    _gerar_e_guardar_dossie(db, ben)
    registrar(db, "creche_dossie_gerado", ator="rh", ator_detalhe=rh.email,
              candidato_id=ben.candidato_id)
    db.commit()
    return {"gerado_em": ben.dossie_gerado_em}


@router.get("/rh/creche/levantamentos/{beneficio_id}/dossie")
def baixar_dossie(beneficio_id: uuid.UUID, db: Session = Depends(get_db)) -> Response:
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    if not ben.dossie_pdf_key:
        _gerar_e_guardar_dossie(db, ben)
        db.commit()
    dados = storage.ler(ben.dossie_pdf_key)
    col = db.get(Candidato, ben.candidato_id)
    nome = (col.nome_completo or "colaborador").replace(" ", "-").lower()
    return Response(content=dados, media_type="application/pdf",
                    headers={"Content-Disposition":
                             f'inline; filename="dossie-creche-{nome}.pdf"'})


_CT_POR_EXT = {
    "pdf": "application/pdf", "jpg": "image/jpeg", "jpeg": "image/jpeg",
    "png": "image/png", "gif": "image/gif", "webp": "image/webp", "heic": "image/heic",
}


@router.get("/rh/creche/levantamentos/{beneficio_id}/crianca/{crianca_id}/documento/{tipo}")
def baixar_doc_crianca(beneficio_id: uuid.UUID, crianca_id: uuid.UUID, tipo: str,
                       db: Session = Depends(get_db),
                       rh: UsuarioRH = Depends(requer_rh)) -> Response:
    """Serve o documento (certidão/guarda) enviado para uma criança, para o RH
    conferir individualmente — aos moldes dos documentos cadastrais."""
    if tipo not in ("certidao", "guarda"):
        raise HTTPException(status_code=422, detail="tipo_invalido")
    c = db.get(CriancaCreche, crianca_id)
    if c is None or c.beneficio_id != beneficio_id:
        raise HTTPException(status_code=404, detail="crianca_nao_encontrada")
    key = c.certidao_key if tipo == "certidao" else c.guarda_key
    if not key:
        raise HTTPException(status_code=404, detail="documento_nao_encontrado")
    # resolve TUDO do banco antes de tocar o storage (evita DetachedInstanceError)
    ext = key.rsplit(".", 1)[-1].lower()
    content_type = _CT_POR_EXT.get(ext, "application/octet-stream")
    nome = f"{tipo}-{c.nome.replace(' ', '-').lower()}.{ext}"
    registrar(db, "creche_doc_crianca_visto", ator="rh", ator_detalhe=rh.email,
              candidato_id=None, detalhe={"beneficio": str(beneficio_id), "tipo": tipo})
    db.commit()
    try:
        dados = storage.ler(key)
    except Exception:
        raise HTTPException(status_code=404, detail="arquivo_nao_encontrado")
    return Response(content=dados, media_type=content_type,
                    headers={"Content-Disposition": f'inline; filename="{nome}"'})


@router.get("/rh/creche/levantamentos/{beneficio_id}/documento/{tipo}")
def previa_documento(beneficio_id: uuid.UUID, tipo: str, db: Session = Depends(get_db)) -> Response:
    """Prévia do requerimento preenchido (tipo=requerimento) ou da declaração
    modelo (tipo=declaracao) no timbrado."""
    from app.services.creche_pdf import (gerar_declaracao_modelo,
                                        gerar_requerimento_creche)
    ben = db.get(BeneficioCreche, beneficio_id)
    if ben is None:
        raise HTTPException(status_code=404, detail="beneficio_nao_encontrado")
    if tipo == "requerimento":
        pdf = gerar_requerimento_creche(db, ben)
    elif tipo == "declaracao":
        pdf = gerar_declaracao_modelo(db, ben)
    else:
        raise HTTPException(status_code=422, detail="tipo_invalido")
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{tipo}.pdf"'})
