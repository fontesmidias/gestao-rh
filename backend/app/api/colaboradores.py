"""Dash de colaboradores do RH: visão com filtros e exportação Excel completa
(linha a linha, com todas as respostas do formulário), importação em massa da
base do Tirvu e controles de vínculo (efetivar, desligar, transferir posto)."""

import io
import re
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.auth_rh import requer_rh
from app.core.db import get_db
from app.models.candidato import Candidato, PostoServico, StatusCandidato
from app.models.ficha import (ContatoEmergencia, DadosPessoais,
                              DadosProfissionaisBancarios, Dependente,
                              DocumentosIdentificacao, Endereco, FichaEmergencia,
                              ValeTransporte)
from app.services.auditoria import registrar
from app.services.export_planilha import linha_completa as _linha_completa
from app.services.export_planilha import montar_workbook
from app.services.idempotencia import travar_por
from app.models.usuario_rh import UsuarioRH

router = APIRouter(tags=["colaboradores-rh"], dependencies=[Depends(requer_rh)])


def _filtrar(db: Session, status: str | None, busca: str | None,
             situacao: str | None = None, posto_id: uuid.UUID | None = None,
             so_colaboradores: bool = True) -> list[Candidato]:
    q = select(Candidato).order_by(Candidato.criado_em.desc())
    # A página Colaboradores mostra APENAS quem já é colaborador de fato: quem
    # foi importado do Tirvu ou efetivado (situacao preenchida). Quem ainda está
    # no fluxo de admissão (situacao NULL) aparece só em Admissões — antes
    # vazava para cá porque candidato e colaborador eram a mesma listagem.
    if so_colaboradores:
        q = q.where(Candidato.situacao.is_not(None))
    if status:
        q = q.where(Candidato.status == StatusCandidato(status))
    if situacao:
        q = q.where(Candidato.situacao == situacao)
    if posto_id:
        q = q.where(Candidato.posto_servico_id == posto_id)
    candidatos = db.scalars(q).all()
    if busca:
        termo = busca.strip().lower()
        so_digitos = "".join(ch for ch in termo if ch.isdigit())
        # CPF agora é campo nativo do Candidato (importação); mas para quem veio
        # da admissão o CPF vive na ficha de documentos. Consulto ambos.
        cpfs = {}
        if so_digitos:
            for doc in db.scalars(select(DocumentosIdentificacao)).all():
                cpfs[doc.candidato_id] = doc.cpf or ""
        candidatos = [
            c for c in candidatos
            # e-mail e celular podem ser None (convite sem e-mail, v1.3) —
            # era isto que derrubava a busca com 500.
            if termo in (c.nome_completo or "").lower()
            or termo in (c.email or "").lower()
            or (so_digitos and (so_digitos in _so_digitos(c.cpf)
                                or so_digitos in cpfs.get(c.id, "")))
        ]
    return candidatos


@router.get("/rh/colaboradores")
def listar(status: str | None = None, busca: str | None = None,
           situacao: str | None = None, posto_id: uuid.UUID | None = None,
           incluir_admissao: bool = False,
           db: Session = Depends(get_db)) -> list[dict]:
    # incluir_admissao=True traz também quem ainda está no fluxo de admissão
    # (para o RH localizar e efetivar um aprovado, por ex.).
    candidatos = _filtrar(db, status, busca, situacao, posto_id,
                          so_colaboradores=not incluir_admissao)
    # nomes dos postos em um só lookup (evita N+1 na lista de 1.156)
    postos = {p.id: p.nome for p in db.scalars(select(PostoServico)).all()}
    saida = []
    for c in candidatos:
        # Para importados, CPF/nascimento já são nativos; para admissão, caem na
        # ficha. Só busco a ficha quando o campo nativo está vazio.
        cpf = c.cpf
        nasc = c.data_nascimento
        if not cpf:
            d = db.get(DocumentosIdentificacao, c.id)
            cpf = d.cpf if d else None
        if not nasc:
            p = db.get(DadosPessoais, c.id)
            nasc = p.data_nascimento if p else None
        saida.append({
            "id": c.id, "nome_completo": c.nome_completo, "email": c.email,
            "celular_whatsapp": c.celular_whatsapp, "status": c.status,
            "situacao": c.situacao, "origem": c.origem,
            "cpf": cpf, "nascimento": nasc, "matricula": c.matricula,
            "posto_id": c.posto_servico_id,
            "posto_nome": postos.get(c.posto_servico_id),
            "data_admissao": c.data_admissao,
            "data_desligamento": c.data_desligamento,
            "na_dominio_em": c.na_dominio_em,
            # indício de que já existe no Tirvu — o front avisa (não bloqueia)
            # antes de reverter a candidato (feedback 2026-07-21).
            "indicio_tirvu": _indicio_tirvu(c),
            # campos vazios do cadastro (importados do Tirvu vêm com buracos) —
            # o RH vê na lista quem precisa completar (feedback 2026-07-21).
            "dados_faltando": _dados_faltando(db, c, cpf, nasc),
            "criado_em": c.criado_em,
            "dossie_gerado_em": c.dossie_gerado_em,
        })
    return saida


@router.get("/rh/colaboradores/exportar")
def exportar(status: str | None = None, busca: str | None = None,
             situacao: str | None = None, posto_id: uuid.UUID | None = None,
             incluir_admissao: bool = False,
             db: Session = Depends(get_db),
             rh: UsuarioRH = Depends(requer_rh)) -> Response:
    """Excel com uma linha por colaborador e TODAS as respostas do formulário.
    Respeita os mesmos filtros da tela (só-colaboradores por padrão)."""
    candidatos = _filtrar(db, status, busca, situacao, posto_id,
                          so_colaboradores=not incluir_admissao)
    conteudo = montar_workbook([_linha_completa(db, c) for c in candidatos])
    registrar(db, "colaboradores_exportados", ator="rh", ator_detalhe=rh.email,
              detalhe={"linhas": len(candidatos), "status": status or "todos"})
    db.commit()
    agora = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=conteudo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="colaboradores-{agora}.xlsx"'},
    )


# ======================================================================
# Exportação para o Tirvu (layout de importação de admissões)
# ======================================================================
#
# Só se exporta COLABORADOR: a pessoa nasce aqui como candidata e só passa a
# existir no Tirvu quando é efetivada (decisão do Bruno, 2026-07-19). Por
# padrão saem apenas os que vieram da admissão — os importados do Tirvu já
# existem lá e seriam ignorados por ele de qualquer forma.


def _colaboradores_para_tirvu(db: Session, status: str | None, busca: str | None,
                              situacao: str | None, posto_id: uuid.UUID | None,
                              ids: str | None, incluir_importados: bool) -> list[Candidato]:
    if ids:
        alvo = [i for i in (x.strip() for x in ids.split(",")) if i]
        return [c for i in alvo if (c := db.get(Candidato, uuid.UUID(i))) is not None]
    candidatos = _filtrar(db, status, busca, situacao, posto_id, so_colaboradores=True)
    if not incluir_importados:
        candidatos = [c for c in candidatos if c.origem != "importacao"]
    return candidatos


# ATENÇÃO: rotas literais ANTES da paramétrica /{cid} (senão vira UUID inválido).
@router.get("/rh/colaboradores/tirvu-pendencias")
def pendencias_tirvu(status: str | None = None, busca: str | None = None,
                     situacao: str | None = None, posto_id: uuid.UUID | None = None,
                     ids: str | None = None, incluir_importados: bool = False,
                     db: Session = Depends(get_db)) -> dict:
    """Pré-checagem do export: o Tirvu RECUSA linha sem CTPS/PIS. O front avisa
    ANTES do download — melhor saber aqui do que descobrir lá."""
    from app.services.export_tirvu import linha_tirvu, pendencias_linha

    candidatos = _colaboradores_para_tirvu(db, status, busca, situacao, posto_id,
                                           ids, incluir_importados)
    problemas = []
    for c in candidatos:
        faltas = pendencias_linha(linha_tirvu(db, c))
        if faltas:
            problemas.append({"id": c.id, "nome": c.nome_completo, "faltam": faltas})
    return {"total": len(candidatos), "com_pendencia": problemas}


@router.get("/rh/colaboradores/exportar-tirvu")
def exportar_tirvu_massa(status: str | None = None, busca: str | None = None,
                         situacao: str | None = None, posto_id: uuid.UUID | None = None,
                         ids: str | None = None, incluir_importados: bool = False,
                         db: Session = Depends(get_db),
                         rh: UsuarioRH = Depends(requer_rh)) -> Response:
    """Planilha no layout de importação de admissões do Tirvu (28 colunas em
    ordem fixa). É o artefato mais sensível do sistema (CPF+PIS+salário em
    massa): auditoria sempre, com quem baixou, quantas linhas e quais postos."""
    from app.services.export_tirvu import linha_tirvu, montar_workbook_tirvu

    candidatos = _colaboradores_para_tirvu(db, status, busca, situacao, posto_id,
                                           ids, incluir_importados)
    if not candidatos:
        raise HTTPException(status_code=404, detail="nenhum_colaborador")
    # gerar_matricula=True: quem não tem matrícula recebe a automática (999+seq),
    # gravada no cadastro (o commit abaixo persiste).
    linhas = [linha_tirvu(db, c, gerar_matricula=True) for c in candidatos]
    # planilha CRUA no formato exato do Tirvu (aba Plan1, sem filtro/cor/freeze)
    conteudo = montar_workbook_tirvu(linhas)
    registrar(db, "tirvu_exportado", ator="rh", ator_detalhe=rh.email,
              detalhe={"linhas": len(linhas), "incluiu_importados": incluir_importados,
                       "postos": sorted({l["Posto de Serviço"] for l in linhas
                                         if l["Posto de Serviço"]})})
    db.commit()
    agora = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=conteudo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="importacao-tirvu-{agora}.xlsx"'})


# ======================================================================
# Importação em massa da base do Tirvu
# ======================================================================
#
# Mapeia cabeçalhos da planilha do Tirvu -> campos fixos do Candidato. Tudo que
# NÃO estiver aqui entra em `dados_tirvu` (colunas dinâmicas). O casamento de
# cabeçalho é tolerante (sem acento, minúsculo, sem espaços duplos).

_MAPA_TIRVU = {
    "cpf": "cpf",
    "colaborador": "nome_completo",
    "nome": "nome_completo",
    "matricula": "matricula",
    "nascimento": "data_nascimento",
    "data de nascimento": "data_nascimento",
    "cargo": "cargo_funcao",
    "lotacao": "_lotacao",           # vira posto (casado/criado à parte)
    "admissao": "data_admissao",
    "demissao": "data_desligamento",
    "status": "_situacao",           # ATIVO/DEMITIDO -> situacao
    "e-mail": "email",
    "email": "email",
    "telefone": "celular_whatsapp",
    "salario": "salario_base",
}


def _norm(txt: str) -> str:
    import unicodedata
    txt = unicodedata.normalize("NFKD", str(txt or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", txt).strip().lower()


def _so_digitos(v) -> str:
    return "".join(ch for ch in str(v or "") if ch.isdigit())


def _cpf_fmt(v) -> str | None:
    d = _so_digitos(v)
    if len(d) != 11:
        return None
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"


def _casar_posto(db: Session, cache: dict, lotacao: str) -> uuid.UUID | None:
    """Casa a 'Lotação' do Tirvu com um posto existente (por nome, case-insens.);
    se não existir, cria um posto novo (nunca falha a linha por isto)."""
    nome = (lotacao or "").strip()
    if not nome:
        return None
    chave = nome.lower()
    if chave in cache:
        return cache[chave]
    posto = db.scalar(select(PostoServico).where(PostoServico.nome.ilike(nome)))
    if posto is None:
        posto = PostoServico(nome=nome)
        db.add(posto)
        db.flush()
    cache[chave] = posto.id
    return posto.id


@router.post("/rh/colaboradores/importar")
async def importar_colaboradores(arquivo: UploadFile,
                                 db: Session = Depends(get_db),
                                 rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Importa a base de colaboradores ativos a partir da planilha (.xlsx) do
    Tirvu. Idempotente por CPF: linha cujo CPF já existe é ATUALIZADA, não
    duplicada. Colunas conhecidas viram campos; as demais entram em dados_tirvu.
    A 'Lotação' é casada com um posto (criado se não existir)."""
    from openpyxl import load_workbook

    try:
        conteudo = await arquivo.read()
    finally:
        # descarta o spool em disco do Starlette — planilha com CPFs não
        # persiste no container depois de processada (regra transversal)
        await arquivo.close()
    try:
        wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="arquivo_invalido")
    ws = wb.active
    linhas = ws.iter_rows(values_only=True)
    try:
        cabecalho = [str(c or "").strip() for c in next(linhas)]
    except StopIteration:
        raise HTTPException(status_code=422, detail="planilha_vazia")

    # posição da coluna de CPF (obrigatória para dedup)
    norm_cab = [_norm(c) for c in cabecalho]
    if "cpf" not in norm_cab:
        raise HTTPException(status_code=422, detail="sem_coluna_cpf")

    # índice de CPFs já cadastrados -> registro
    ja = {c.cpf: c for c in db.scalars(
        select(Candidato).where(Candidato.cpf.is_not(None))).all()}
    cache_postos: dict = {}
    criados = atualizados = ignorados = 0
    sem_cpf = 0

    for bruta in linhas:
        if bruta is None or all(v in (None, "") for v in bruta):
            continue
        valores = list(bruta) + [None] * (len(cabecalho) - len(bruta))
        cpf = _cpf_fmt(valores[norm_cab.index("cpf")])
        if not cpf:
            sem_cpf += 1
            continue

        campos: dict = {}
        lotacao = None
        situacao_bruta = None
        dinamicos: dict = {}
        for titulo, nrm, valor in zip(cabecalho, norm_cab, valores):
            destino = _MAPA_TIRVU.get(nrm)
            v = "" if valor is None else str(valor).strip()
            if destino == "_lotacao":
                lotacao = v
            elif destino == "_situacao":
                situacao_bruta = v
            elif destino == "cpf":
                pass  # já tratado
            elif destino:
                campos[destino] = v or None
            elif v:
                dinamicos[titulo] = v

        situacao = "desligado" if _norm(situacao_bruta).startswith(("demit", "inativ", "deslig")) \
            else "ativo"

        alvo = ja.get(cpf)
        if alvo is None:
            alvo = Candidato(cpf=cpf, nome_completo=campos.get("nome_completo") or "(sem nome)",
                             origem="importacao")
            db.add(alvo)
            ja[cpf] = alvo
            criados += 1
        else:
            atualizados += 1
        # aplica campos fixos. NÃO sobrescreve com vazio o nome nem a MATRÍCULA:
        # a matrícula (999NNNN gerada ou a real do Tirvu) é estável e serve de
        # âncora — uma célula "matricula" em branco na planilha zeraria o
        # cadastro. O Tirvu reimporta e ATUALIZA a matrícula quando ela vem
        # preenchida (o setattr abaixo faz isso); só protege contra o vazio.
        for k, val in campos.items():
            if k in ("nome_completo", "matricula") and not val:
                continue
            setattr(alvo, k, val)
        alvo.situacao = situacao  # vínculo (ativo/desligado) vem do Tirvu
        # `status` é fluxo: importado nunca passou pelo funil daqui (v1.69).
        alvo.status = StatusCandidato.importado
        if lotacao:
            alvo.posto_servico_id = _casar_posto(db, cache_postos, lotacao)
        # mescla dinâmicos preservando o que já houver
        base = dict(alvo.dados_tirvu or {})
        base.update(dinamicos)
        alvo.dados_tirvu = base

    registrar(db, "colaboradores_importados", ator="rh", ator_detalhe=rh.email,
              detalhe={"criados": criados, "atualizados": atualizados,
                       "sem_cpf": sem_cpf, "postos_novos": len(cache_postos)})
    db.commit()
    return {"criados": criados, "atualizados": atualizados,
            "sem_cpf": sem_cpf, "postos_tocados": len(cache_postos),
            "total_base": len(ja)}


# ======================================================================
# Controles de vínculo: efetivar candidato, desligar, transferir posto
# ======================================================================


class DesligamentoIn(BaseModel):
    data_desligamento: str  # dd/mm/aaaa


class TransferenciaIn(BaseModel):
    posto_id: uuid.UUID
    data_transferencia: str | None = None  # registrada em dados_tirvu (histórico)


def _get_colab(db: Session, cid: uuid.UUID) -> Candidato:
    c = db.get(Candidato, cid)
    if c is None:
        raise HTTPException(status_code=404, detail="colaborador_nao_encontrado")
    return c


def _efetivar_um(db: Session, c: Candidato) -> None:
    c.situacao = "ativo"  # vínculo
    # `status` é fluxo: efetivar aqui = admissão aprovada e concluída (v1.69).
    c.status = StatusCandidato.aprovado
    if not c.data_admissao:
        c.data_admissao = datetime.now(timezone.utc).strftime("%d/%m/%Y")


# Reverter colaborador -> candidato (feedback 2026-07-21: "converti por engano
# e não consigo voltar"). Destinos válidos = fases do fluxo em que faz sentido
# reentrar (não os estados terminais nem o vínculo ativo/desligado).
DESTINOS_REVERTER = {
    StatusCandidato.convidado,
    StatusCandidato.em_revisao,
}


def _dados_faltando(db: Session, c: Candidato, cpf: str | None,
                    nascimento: str | None) -> list[str]:
    """Campos vazios no cadastro que travam a operação do RH (feedback
    2026-07-21: "tem gente que foi importada do Tirvu e tem dados em branco").
    Checa o que é preciso para CONTATAR e para EFETIVAR/EXPORTAR — não a ficha
    inteira (o importado nunca preencheu ficha e não deve parecer 'incompleto'
    por causa disso)."""
    faltas = []
    if not cpf:
        faltas.append("CPF")
    if not nascimento:
        faltas.append("nascimento")
    if not c.email:
        faltas.append("e-mail")
    if not c.celular_whatsapp:
        faltas.append("telefone")
    if not c.posto_servico_id:
        faltas.append("posto")
    if not c.cargo_funcao:
        faltas.append("cargo")
    if not c.jornada_id:
        faltas.append("jornada")
    return faltas


def _indicio_tirvu(c: Candidato) -> str | None:
    """Sinaliza que o colaborador provavelmente já existe no Tirvu — para o RH
    ser AVISADO (nunca bloqueado, decisão do Bruno 2026-07-21) antes de reverter.
    Dois indícios: veio do Tirvu (importação) ou já ganhou matrícula 999NNNN
    (gerada só no export de admissões para o Tirvu)."""
    if c.origem == "importacao":
        return "importado do Tirvu"
    if (c.matricula or "").startswith("999"):
        return "já teve matrícula gerada para o Tirvu"
    return None


def _reverter_um(c: Candidato, destino: StatusCandidato) -> None:
    """Zera o vínculo de colaborador e devolve o registro ao fluxo de admissão.
    Preserva a matrícula (reusada se a pessoa voltar a ser efetivada) e os dados;
    só limpa o que caracteriza o vínculo ativo."""
    c.situacao = None
    c.status = destino
    c.data_admissao = None
    c.data_desligamento = None


class LoteEfetivarIn(BaseModel):
    ids: list[uuid.UUID]


# ATENÇÃO: a rota específica /lote/efetivar precisa vir ANTES da paramétrica
# /{cid}/efetivar, senão "lote" é interpretado como um UUID inválido (422).
@router.post("/rh/colaboradores/lote/efetivar")
def efetivar_lote(payload: LoteEfetivarIn, db: Session = Depends(get_db),
                  rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Efetiva vários candidatos de uma vez. Já-colaboradores são pulados."""
    efetivados, pulados = 0, 0
    for cid in payload.ids:
        c = db.get(Candidato, cid)
        if c is None:
            continue
        if c.situacao:  # já é colaborador (ativo/desligado): não mexe
            pulados += 1
            continue
        _efetivar_um(db, c)
        efetivados += 1
    registrar(db, "colaboradores_efetivados_lote", ator="rh", ator_detalhe=rh.email,
              detalhe={"efetivados": efetivados, "pulados": pulados})
    db.commit()
    return {"efetivados": efetivados, "pulados": pulados}


class AcaoMassaColabIn(BaseModel):
    ids: list[uuid.UUID]
    acao: str  # "desligar" | "reativar" | "marcar_dominio" | "desmarcar_dominio"
    data_desligamento: str | None = None  # obrigatória para "desligar"


@router.post("/rh/colaboradores/lote/acao")
def acao_massa_colaboradores(payload: AcaoMassaColabIn, db: Session = Depends(get_db),
                             rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Ação em massa nos colaboradores selecionados: desligar (com data) ou
    reativar. Não há exclusão — registro trabalhista não se apaga; desligar é o
    correto. Só age sobre quem já é colaborador (tem situação)."""
    if payload.acao not in ("desligar", "reativar", "marcar_dominio", "desmarcar_dominio"):
        raise HTTPException(status_code=422, detail="acao_invalida")
    if payload.acao == "desligar" and not (payload.data_desligamento or "").strip():
        raise HTTPException(status_code=422, detail="data_desligamento_obrigatoria")
    afetados, pulados = 0, 0
    for cid in payload.ids:
        c = db.get(Candidato, cid)
        if c is None:
            pulados += 1
            continue
        # conciliação com a Domínio vale para qualquer admissão (mesmo em curso)
        if payload.acao in ("marcar_dominio", "desmarcar_dominio"):
            c.na_dominio_em = (datetime.now(timezone.utc)
                               if payload.acao == "marcar_dominio" else None)
            afetados += 1
            continue
        if not c.situacao:  # desligar/reativar: ignora candidatos em admissão
            pulados += 1
            continue
        if payload.acao == "desligar":
            c.situacao = "desligado"  # vínculo; `status` (fluxo) não muda
            c.data_desligamento = payload.data_desligamento.strip()
            from app.api.creche import encerrar_creche_no_desligamento
            encerrar_creche_no_desligamento(db, c.id)  # encerra benefício ativo
        else:  # reativar
            c.situacao = "ativo"
            c.data_desligamento = None
        afetados += 1
    registrar(db, "colaboradores_acao_massa", ator="rh", ator_detalhe=rh.email,
              detalhe={"acao": payload.acao, "afetados": afetados, "pulados": pulados})
    db.commit()
    return {"afetados": afetados, "pulados": pulados}


class ReverterIn(BaseModel):
    destino: str = "convidado"      # "convidado" | "em_revisao"
    motivo: str                     # obrigatório (auditoria)


class LoteReverterIn(BaseModel):
    ids: list[uuid.UUID]
    destino: str = "convidado"
    motivo: str


def _validar_reverter(destino: str, motivo: str) -> StatusCandidato:
    if not (motivo or "").strip():
        raise HTTPException(status_code=422, detail="motivo_obrigatorio")
    try:
        alvo = StatusCandidato(destino)
    except ValueError:
        raise HTTPException(status_code=422, detail="destino_invalido")
    if alvo not in DESTINOS_REVERTER:
        raise HTTPException(status_code=422, detail="destino_invalido")
    return alvo


# Específica ANTES da paramétrica /{cid}/... (senão "lote" vira UUID inválido).
@router.post("/rh/colaboradores/lote/reverter")
def reverter_lote(payload: LoteReverterIn, db: Session = Depends(get_db),
                  rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Reverte vários colaboradores a candidatos de uma vez. Só age sobre quem
    JÁ é colaborador (tem situação); candidatos em admissão são pulados. O aviso
    de indício-Tirvu é responsabilidade do front (nunca bloqueia — decisão do
    Bruno); aqui só executa e audita."""
    alvo = _validar_reverter(payload.destino, payload.motivo)
    revertidos, pulados = 0, 0
    for cid in payload.ids:
        c = db.get(Candidato, cid)
        if c is None or not c.situacao:  # inexistente ou já em admissão
            pulados += 1
            continue
        _reverter_um(c, alvo)
        registrar(db, "colaborador_revertido", ator="rh", ator_detalhe=rh.email,
                  candidato_id=c.id,
                  detalhe={"nome": c.nome_completo, "destino": alvo.value,
                           "motivo": payload.motivo.strip(),
                           "tirvu": _indicio_tirvu(c)})
        revertidos += 1
    registrar(db, "colaboradores_revertidos_lote", ator="rh", ator_detalhe=rh.email,
              detalhe={"revertidos": revertidos, "pulados": pulados,
                       "destino": alvo.value})
    db.commit()
    return {"revertidos": revertidos, "pulados": pulados}


@router.post("/rh/colaboradores/{cid}/reverter")
def reverter(cid: uuid.UUID, payload: ReverterIn, db: Session = Depends(get_db),
             rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Reverte um colaborador a candidato (converteu por engano, ou precisa
    reprocessar). Preserva matrícula e dados; motivo é obrigatório."""
    alvo = _validar_reverter(payload.destino, payload.motivo)
    c = _get_colab(db, cid)
    if not c.situacao:
        raise HTTPException(status_code=422, detail="nao_e_colaborador")
    _reverter_um(c, alvo)
    registrar(db, "colaborador_revertido", ator="rh", ator_detalhe=rh.email,
              candidato_id=c.id,
              detalhe={"nome": c.nome_completo, "destino": alvo.value,
                       "motivo": payload.motivo.strip(), "tirvu": _indicio_tirvu(c)})
    db.commit()
    return {"id": c.id, "situacao": c.situacao, "status": c.status.value}


@router.post("/rh/colaboradores/{cid}/efetivar",
             dependencies=[Depends(travar_por("efetivar"))])
def efetivar(cid: uuid.UUID, db: Session = Depends(get_db),
             rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Transforma um candidato aprovado em colaborador ativo (mesmo registro)."""
    c = _get_colab(db, cid)
    _efetivar_um(db, c)
    registrar(db, "colaborador_efetivado", ator="rh", ator_detalhe=rh.email,
              candidato_id=c.id, detalhe={"nome": c.nome_completo})
    db.commit()
    return {"id": c.id, "situacao": c.situacao, "status": c.status,
            "data_admissao": c.data_admissao}


@router.post("/rh/colaboradores/{cid}/desligar")
def desligar(cid: uuid.UUID, payload: DesligamentoIn, db: Session = Depends(get_db),
             rh: UsuarioRH = Depends(requer_rh)) -> dict:
    c = _get_colab(db, cid)
    c.situacao = "desligado"  # vínculo; `status` (fluxo) não muda
    c.data_desligamento = payload.data_desligamento.strip() or None
    # encerra o benefício creche ativo (não se reembolsa quem saiu)
    from app.api.creche import encerrar_creche_no_desligamento
    encerrar_creche_no_desligamento(db, c.id)
    registrar(db, "colaborador_desligado", ator="rh", ator_detalhe=rh.email,
              candidato_id=c.id,
              detalhe={"nome": c.nome_completo, "data": c.data_desligamento})
    db.commit()
    return {"id": c.id, "situacao": c.situacao, "data_desligamento": c.data_desligamento}


@router.post("/rh/colaboradores/{cid}/transferir")
def transferir(cid: uuid.UUID, payload: TransferenciaIn, db: Session = Depends(get_db),
               rh: UsuarioRH = Depends(requer_rh)) -> dict:
    c = _get_colab(db, cid)
    posto = db.get(PostoServico, payload.posto_id)
    if posto is None:
        raise HTTPException(status_code=404, detail="posto_nao_encontrado")
    origem = str(c.posto_servico_id) if c.posto_servico_id else None
    c.posto_servico_id = posto.id
    # histórico de transferências guardado nos dados dinâmicos (sem nova tabela)
    hist = dict(c.dados_tirvu or {})
    linha = f"{payload.data_transferencia or datetime.now(timezone.utc).strftime('%d/%m/%Y')} -> {posto.nome}"
    hist["Transferências"] = (hist.get("Transferências", "") + "; " + linha).strip("; ")
    c.dados_tirvu = hist
    registrar(db, "colaborador_transferido", ator="rh", ator_detalhe=rh.email,
              candidato_id=c.id,
              detalhe={"nome": c.nome_completo, "de": origem, "para": str(posto.id)})
    db.commit()
    return {"id": c.id, "posto_servico_id": c.posto_servico_id, "posto_nome": posto.nome}


# ======================================================================
# Vínculo em massa: posto, cargo e jornada conforme o Tirvu (v2.39)
# ======================================================================
#
# Pedido do Bruno (2026-08-01): "precisa vincular os colaboradores em massa
# também a seus respectivos postos, cargos e jornadas, conforme Tirvu, quero
# evitar trabalho manual".
#
# Por que NÃO virou parte da importação de colaboradores, que lê a mesma
# planilha: aquela grava direto, e aqui o RH precisa CONFERIR antes — são ~1.000
# registros e o que se sobrescreve não volta. Preview e aplicação separados,
# como na Incidência de Benefícios.


def _mapas_para_vinculo(db: Session) -> tuple[dict, dict, dict]:
    """Tudo que o cruzamento precisa, em 3 consultas — nunca uma por linha.

    Com 1.156 pessoas na planilha, consultar por linha seria a diferença entre
    segundos e minutos (e a lição de N+1 que a v2.15 já cobrou neste projeto).
    """
    from app.models.candidato import Jornada
    from app.services.vinculo_tirvu import normalizar, so_digitos

    pessoas = {}
    for c in db.scalars(select(Candidato).where(Candidato.cpf.is_not(None))).all():
        pessoas[so_digitos(c.cpf)] = c
    # PCD mora na FICHA (`DadosPessoais`, 1:1), não no Candidato. Carrego em
    # lote e penduro no objeto como atributo efêmero, só para o analisador ler:
    # é o que evita uma consulta de ficha por linha da planilha.
    pcd_por_candidato = {dp.candidato_id: dp.pcd for dp in db.scalars(select(DadosPessoais)).all()}
    for c in pessoas.values():
        c.pcd = pcd_por_candidato.get(c.id)
    jornadas = {normalizar(j.descricao): j.id for j in db.scalars(select(Jornada)).all()}
    postos = {normalizar(p.nome): p.id for p in db.scalars(select(PostoServico)).all()}
    # O de-para confirmado pelo RH entra no MESMO mapa (v2.40): é o que faz
    # "INEP ADM" casar com o posto certo. Vem DEPOIS do nome, e por isso vence
    # — foi decisão humana explícita, enquanto o casamento por nome é
    # coincidência de texto.
    from app.models.candidato import LotacaoTirvu
    for m in db.scalars(select(LotacaoTirvu)).all():
        postos[m.lotacao_normalizada] = m.posto_servico_id
    return pessoas, jornadas, postos


def _resumo_decisao(d) -> dict:
    return {
        "cpf": d.cpf, "nome": d.nome, "achou_cadastro": d.achou_cadastro,
        "jornada": {"texto": d.jornada_texto, "id": d.jornada_id,
                    "situacao": d.jornada_situacao},
        "cargo": {"texto": d.cargo_texto, "situacao": d.cargo_situacao,
                  "atual": d.cargo_atual},
        "posto": {"texto": d.lotacao_texto, "id": d.posto_id,
                  "situacao": d.posto_situacao},
        "pcd": {"valor": d.pcd, "deficiencia": d.pcd_deficiencia,
                "situacao": d.pcd_situacao},
    }


@router.post("/rh/colaboradores/vinculos/preview")
async def preview_vinculos(arquivo: UploadFile, db: Session = Depends(get_db),
                           _rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Lê a planilha de Colaboradores do Tirvu e PROPÕE os vínculos. Não grava.

    Devolve o que está pronto (campo vazio no portal), o que DIVERGE (valor
    diferente aqui — pode ser correção feita à mão, então é decisão humana) e o
    que não tem par na base, com quantas pessoas dependem de cada item.
    """
    from app.api.postos import _ler_linhas_xlsx
    from app.services.vinculo_tirvu import analisar

    try:
        linhas = _ler_linhas_xlsx(await arquivo.read())
    except Exception as exc:
        raise HTTPException(status_code=422, detail="planilha_ilegivel") from exc
    finally:
        await arquivo.close()   # spool em disco com CPF de mil pessoas

    pessoas, jornadas, postos = _mapas_para_vinculo(db)
    try:
        analise = analisar(linhas, candidatos_por_cpf=pessoas,
                           jornadas_por_descricao=jornadas, postos_por_nome=postos)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    def fila(d: dict) -> list[dict]:
        return [{"texto": k, "pessoas": v}
                for k, v in sorted(d.items(), key=lambda kv: -kv[1])]

    return {
        "linhas": len(analise.decisoes),
        "sem_cpf": analise.sem_cpf,
        "fora_da_base": len(analise.fora_da_base),
        "prontas": len(analise.prontas),
        "divergentes": len(analise.divergentes),
        "jornadas_sem_par": fila(analise.jornadas_sem_par),
        "lotacoes_sem_par": fila(analise.lotacoes_sem_par),
        # O que a tela precisa para o RH conferir e para o aplicar receber de volta
        "itens": [_resumo_decisao(d) for d in analise.prontas],
        "itens_divergentes": [_resumo_decisao(d) for d in analise.divergentes],
    }


class VinculoItemIn(BaseModel):
    cpf: str
    jornada_id: uuid.UUID | None = None
    cargo_funcao: str | None = None
    posto_id: uuid.UUID | None = None
    pcd: bool | None = None
    pcd_deficiencia: str | None = None


class AplicarVinculosIn(BaseModel):
    itens: list[VinculoItemIn]


@router.post("/rh/colaboradores/vinculos/aplicar")
def aplicar_vinculos(payload: AplicarVinculosIn, db: Session = Depends(get_db),
                     rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Grava os vínculos que o RH confirmou. Só toca o que veio na lista.

    O PCD é dado de saúde e vem da base do Tirvu, não de declaração da pessoa:
    é gravado com o rastro de que veio de lá (auditoria), como qualquer outro
    registro que o RH faz sobre alguém.
    """
    from app.services.vinculo_tirvu import so_digitos

    pessoas = {so_digitos(c.cpf): c
               for c in db.scalars(select(Candidato).where(Candidato.cpf.is_not(None))).all()}
    tocados = {"jornada": 0, "cargo": 0, "posto": 0, "pcd": 0}
    ignorados = 0
    for item in payload.itens:
        c = pessoas.get(so_digitos(item.cpf))
        if c is None:
            ignorados += 1
            continue
        if item.jornada_id is not None:
            c.jornada_id = item.jornada_id
            tocados["jornada"] += 1
        if item.cargo_funcao:
            c.cargo_funcao = item.cargo_funcao.strip()
            tocados["cargo"] += 1
        if item.posto_id is not None:
            c.posto_servico_id = item.posto_id
            tocados["posto"] += 1
        if item.pcd is not None:
            dp = db.get(DadosPessoais, c.id)
            if dp is None:
                dp = DadosPessoais(candidato_id=c.id)
                db.add(dp)
            dp.pcd = item.pcd
            if item.pcd_deficiencia and not dp.pcd_tipo:
                dp.pcd_tipo = item.pcd_deficiencia.strip()[:30]
            tocados["pcd"] += 1

    registrar(db, "colaboradores_vinculados_em_massa", ator="rh", ator_detalhe=rh.email,
              detalhe={**tocados, "enviados": len(payload.itens), "ignorados": ignorados,
                       "origem": "planilha do Tirvu"})
    db.commit()
    return {**tocados, "ignorados": ignorados}


# ======================================================================
# De-para LOTAÇÃO → posto de serviço (v2.40)
# ======================================================================
#
# A lotação da planilha vem abreviada ("INEP ADM", "ANAC") e o apelido do posto
# é o padrão longo ("ANAC - 14/2026 - AEROPORTO"). Medido nos dados reais: 11%
# casam sozinhos, e "ANAC" pode ser SEDE ou AEROPORTO — ambiguidade do dado,
# que nenhum algoritmo resolve. Então o sistema ordena candidatos e o RH
# decide; a escolha fica gravada e as importações seguintes não perguntam de
# novo. Mesma mecânica da Incidência de Benefícios, e pela mesma razão.


def _similaridade(a: str, b: str) -> float:
    from difflib import SequenceMatcher

    from app.services.vinculo_tirvu import normalizar
    return SequenceMatcher(None, normalizar(a), normalizar(b)).ratio()


def _sugerir_postos(lotacao: str, postos: list) -> list[dict]:
    """Até 3 postos candidatos, do mais parecido para o menos. NUNCA decide.

    O bônus de contenção é o que resolve o caso real: "ANAC" está contido em
    "ANAC - 14/2026 - AEROPORTO" e em "- SEDE", então os DOIS sobem juntos — é
    exatamente a ambiguidade que precisa chegar aos olhos do RH, e não ser
    desempatada no escuro por um centésimo de similaridade.
    """
    from app.services.vinculo_tirvu import normalizar

    alvo = normalizar(lotacao)
    palavras_alvo = [p for p in alvo.split() if len(p) > 1]
    ranqueados = []
    for p in postos:
        campos = [p.nome or "", p.sigla or "", p.razao_social or ""]
        score = max((_similaridade(lotacao, c) for c in campos if c), default=0.0)
        if alvo and any(alvo in normalizar(c) for c in campos if c):
            score = max(score, 0.9)
        # Semelhança de LETRAS engana, e engana no caso que mais importa: nos
        # dados reais "INEP ADM" (174 pessoas) pontuava 0.67 com "IPAM" e só
        # 0.47 com "INEP - 37/2025 - APOIO ADM", porque as letras I-P-A-M estão
        # todas lá na ordem. Um RH apressado aceitaria a sugestão errada e 174
        # pessoas iriam para o posto de outro contrato.
        #
        # A palavra INTEIRA é o sinal forte: "INEP" aparece como palavra no
        # nome do posto certo e não aparece em "IPAM". Cada palavra do texto do
        # Tirvu que reaparece inteira no posto vale mais que qualquer
        # coincidência de caracteres.
        palavras_posto = {w for c in campos for w in normalizar(c).split()}
        if palavras_alvo:
            casadas = sum(1 for w in palavras_alvo if w in palavras_posto)
            if casadas:
                score = max(score, 0.55 + 0.35 * (casadas / len(palavras_alvo)))
        ranqueados.append((score, p))
    ranqueados.sort(key=lambda t: (-t[0], t[1].nome or ""))
    return [{"posto_id": str(p.id), "posto_nome": p.nome, "sigla": p.sigla,
             "razao_social": p.razao_social, "score": round(s, 2)}
            for s, p in ranqueados[:3] if s > 0.35]


@router.get("/rh/postos/de-para")
def listar_de_para_lotacao(db: Session = Depends(get_db)) -> list[dict]:
    """O que já foi decidido — para o RH conferir e corrigir sem reimportar."""
    from app.models.candidato import LotacaoTirvu

    postos = {p.id: p.nome for p in db.scalars(select(PostoServico)).all()}
    return [{"id": m.id, "lotacao": m.lotacao_rotulo,
             "posto_id": m.posto_servico_id,
             "posto_nome": postos.get(m.posto_servico_id),
             "confirmado_por": m.confirmado_por, "criado_em": m.criado_em}
            for m in db.scalars(select(LotacaoTirvu)
                                .order_by(LotacaoTirvu.lotacao_rotulo)).all()]


@router.post("/rh/postos/de-para/preview")
async def preview_de_para_lotacao(arquivo: UploadFile, db: Session = Depends(get_db),
                                  _rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Lotações da planilha que ainda não têm posto, com candidatos ordenados.

    Ordena pela quantidade de PESSOAS afetadas, não alfabeticamente: resolver
    "INEP ADM" (174 pessoas) antes de uma lotação com 1 vale mais o tempo do
    RH — a lição do módulo de duplicidade de jornada, onde a fila cheia de
    ruído fazia perder o que importava.
    """
    from app.api.postos import _ler_linhas_xlsx
    from app.models.candidato import LotacaoTirvu
    from app.services.vinculo_tirvu import analisar, normalizar

    try:
        linhas = _ler_linhas_xlsx(await arquivo.read())
    except Exception as exc:
        raise HTTPException(status_code=422, detail="planilha_ilegivel") from exc
    finally:
        await arquivo.close()

    pessoas, jornadas, postos_map = _mapas_para_vinculo(db)
    ja = {m.lotacao_normalizada for m in db.scalars(select(LotacaoTirvu)).all()}
    try:
        analise = analisar(linhas, candidatos_por_cpf=pessoas,
                           jornadas_por_descricao=jornadas, postos_por_nome=postos_map)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    postos = db.scalars(select(PostoServico).order_by(PostoServico.nome)).all()
    pendentes = []
    for texto, quantas in sorted(analise.lotacoes_sem_par.items(), key=lambda kv: -kv[1]):
        if normalizar(texto) in ja:
            continue     # já decidido numa rodada anterior
        pendentes.append({"lotacao": texto, "pessoas": quantas,
                          "sugestoes": _sugerir_postos(texto, postos)})
    return {"pendentes": pendentes,
            "total_pessoas": sum(p["pessoas"] for p in pendentes),
            "ja_mapeadas": len(ja),
            "postos": [{"id": str(p.id), "nome": p.nome} for p in postos]}


class DeParaLotacaoItem(BaseModel):
    lotacao: str
    posto_id: uuid.UUID


class DeParaLotacaoIn(BaseModel):
    itens: list[DeParaLotacaoItem]


@router.post("/rh/postos/de-para/confirmar")
def confirmar_de_para_lotacao(payload: DeParaLotacaoIn, db: Session = Depends(get_db),
                              rh: UsuarioRH = Depends(requer_rh)) -> dict:
    """Grava as escolhas do RH. Reenviar a mesma lotação ATUALIZA o destino —
    corrigir um de-para errado não pode exigir apagar e recriar."""
    from app.models.candidato import LotacaoTirvu
    from app.services.vinculo_tirvu import normalizar

    existentes = {m.lotacao_normalizada: m for m in db.scalars(select(LotacaoTirvu)).all()}
    criados = atualizados = 0
    for item in payload.itens:
        chave = normalizar(item.lotacao)
        if not chave:
            continue
        if db.get(PostoServico, item.posto_id) is None:
            raise HTTPException(status_code=422, detail="posto_nao_encontrado")
        m = existentes.get(chave)
        if m:
            if m.posto_servico_id != item.posto_id:
                m.posto_servico_id = item.posto_id
                m.confirmado_por = rh.email
                atualizados += 1
        else:
            db.add(LotacaoTirvu(lotacao_normalizada=chave,
                                lotacao_rotulo=item.lotacao.strip()[:200],
                                posto_servico_id=item.posto_id,
                                confirmado_por=rh.email))
            criados += 1
    registrar(db, "lotacoes_mapeadas", ator="rh", ator_detalhe=rh.email,
              detalhe={"criados": criados, "atualizados": atualizados})
    db.commit()
    return {"criados": criados, "atualizados": atualizados}
